import sqlite3
import re
import unicodedata
import calendar as pycalendar
from datetime import datetime
import customtkinter as ctk
from tkinter import messagebox, ttk
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
import os
import subprocess
import sys
import shutil
import json
from urllib.parse import quote, urlencode
from urllib.request import urlopen, Request

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

APP_SUPPORT_DIR = os.path.join(os.path.expanduser("~/Documents"), "Enmarcados PF Data")


def cargar_env_simple(path):
    if not os.path.exists(path):
        return
    try:
        with open(path, "r", encoding="utf-8") as fh:
            for linea in fh:
                linea = linea.strip()
                if not linea or linea.startswith("#") or "=" not in linea:
                    continue
                clave, valor = linea.split("=", 1)
                os.environ.setdefault(clave.strip(), valor.strip())
    except Exception:
        pass


cargar_env_simple(os.path.join(os.getcwd(), ".env.cloud"))


def ruta_recurso(nombre):
    if getattr(sys, "frozen", False):
        base = getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
        return os.path.join(base, nombre)
    return os.path.join(os.getcwd(), nombre)


def carpeta_datos_app():
    if getattr(sys, "frozen", False):
        os.makedirs(APP_SUPPORT_DIR, exist_ok=True)
        return APP_SUPPORT_DIR
    return os.getcwd()


DB_NAME = os.path.join(carpeta_datos_app(), "facturacion.db")
COLOR_PRINCIPAL = "#57C7C3"
COLOR_PRINCIPAL_HOVER = "#3eb5b1"
COLOR_BORDE = "#D9E3E6"
COLOR_TEXTO_SUAVE = "#5E6B73"
COLOR_PANEL = "#F7FBFB"
LOGO_NEGOCIO = ruta_recurso("5e889a71-beea-42c1-909b-a7bf536098e8.JPG")
NEGOCIO_DIRECCION = "Calle. Juan B. 6, Santo Domingo"
NEGOCIO_TELEFONO = "809 569 2000"
NEGOCIO_RNC = "132-23323-9"
NEGOCIO_INSTAGRAM = "@enmarcadospfrd"
REFERENCIA_COBRO_AUTO = "Abono factura"
DATA_MODE = os.getenv("DATA_MODE", "local").strip().lower()
API_BASE_URL = os.getenv("API_BASE_URL", "http://127.0.0.1:8000").rstrip("/")


# =========================
# UTILIDADES
# =========================
def limpiar_nombre_archivo(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    texto = re.sub(r"[^\w\s-]", "", texto).strip()
    texto = re.sub(r"[-\s]+", "_", texto)
    return texto[:60] if texto else "SIN_NOMBRE"


def usar_backend_nube():
    return DATA_MODE == "cloud" and bool(API_BASE_URL)


def api_get_json(path, params=None):
    if not usar_backend_nube():
        return None
    query = f"?{urlencode(params)}" if params else ""
    url = f"{API_BASE_URL}{path}{query}"
    req = Request(url, headers={"Accept": "application/json"})
    with urlopen(req, timeout=8) as resp:
        return json.loads(resp.read().decode("utf-8"))


def api_send_json(method, path, payload):
    if not usar_backend_nube():
        return None
    url = f"{API_BASE_URL}{path}"
    data = json.dumps(payload).encode("utf-8")
    req = Request(url, data=data, headers={"Content-Type": "application/json", "Accept": "application/json"}, method=method)
    with urlopen(req, timeout=12) as resp:
        return json.loads(resp.read().decode("utf-8"))


def api_post_json(path, payload):
    return api_send_json("POST", path, payload)


def api_put_json(path, payload):
    return api_send_json("PUT", path, payload)


def api_delete_json(path):
    if not usar_backend_nube():
        return None
    url = f"{API_BASE_URL}{path}"
    req = Request(url, method="DELETE", headers={"Accept": "application/json"})
    with urlopen(req, timeout=12) as resp:
        return json.loads(resp.read().decode("utf-8"))


def fecha_hoy_str():
    return datetime.now().strftime("%Y-%m-%d")


def fecha_mes_actual():
    ahora = datetime.now()
    return ahora.year, ahora.month


def filtrar_nombres_clientes(texto, limite=25):
    nombres = [fila[1] for fila in obtener_todos_clientes()]
    texto = (texto or "").strip().lower()
    if not texto:
        return nombres[:limite]

    empiezan = [n for n in nombres if n.lower().startswith(texto)]
    contienen = [n for n in nombres if texto in n.lower() and n not in empiezan]
    return (empiezan + contienen)[:limite]


def carpeta_documentos():
    ruta = os.path.join(carpeta_datos_app(), "documentos_generados")
    os.makedirs(ruta, exist_ok=True)
    return ruta


def carpeta_backup_facturas():
    ruta = os.path.join(carpeta_datos_app(), "facturaspf realizadas")
    os.makedirs(ruta, exist_ok=True)
    return ruta


def carpeta_backups():
    ruta = os.path.join(carpeta_datos_app(), "backup")
    os.makedirs(ruta, exist_ok=True)
    return ruta


def carpeta_backups_nube():
    ruta = os.path.join(carpeta_datos_app(), "backup_nube")
    os.makedirs(ruta, exist_ok=True)
    return ruta


def guardar_copia_backup(ruta_archivo):
    if not ruta_archivo or not os.path.exists(ruta_archivo):
        return None
    destino = os.path.join(carpeta_backup_facturas(), os.path.basename(ruta_archivo))
    shutil.copy2(ruta_archivo, destino)
    return destino


def crear_backup_y_reiniciar_facturacion():
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    carpeta_base = carpeta_backups_nube() if usar_backend_nube() else carpeta_backups()
    carpeta_destino = os.path.join(carpeta_base, timestamp)
    pdfs_generados_destino = os.path.join(carpeta_destino, "documentos_generados")
    pdfs_respaldo_destino = os.path.join(carpeta_destino, "facturaspf realizadas")
    os.makedirs(pdfs_generados_destino, exist_ok=True)
    os.makedirs(pdfs_respaldo_destino, exist_ok=True)

    # Copia completa de la base antes de limpiar.
    if os.path.exists(DB_NAME):
        shutil.copy2(DB_NAME, os.path.join(carpeta_destino, DB_NAME))

    respaldo_nube = None
    if usar_backend_nube():
        respaldo_nube = {
            "exportado_en": datetime.now().isoformat(timespec="seconds"),
            "api_base_url": API_BASE_URL,
            "clientes": api_get_json("/clientes") or [],
            "tarifas": api_get_json("/tarifas") or [],
            "documentos": api_get_json("/documentos", params={"limit": 5000}) or [],
        }
        with open(os.path.join(carpeta_destino, "respaldo_nube.json"), "w", encoding="utf-8") as fh:
            json.dump(respaldo_nube, fh, ensure_ascii=False, indent=2)

    def mover_archivos(origen, destino):
        movidos = 0
        if not os.path.isdir(origen):
            return movidos
        for nombre in os.listdir(origen):
            ruta_origen = os.path.join(origen, nombre)
            if not os.path.isfile(ruta_origen):
                continue
            base, ext = os.path.splitext(nombre)
            ruta_destino = os.path.join(destino, nombre)
            contador = 1
            while os.path.exists(ruta_destino):
                ruta_destino = os.path.join(destino, f"{base}_{contador}{ext}")
                contador += 1
            shutil.move(ruta_origen, ruta_destino)
            movidos += 1
        return movidos

    movidos_generados = mover_archivos(carpeta_documentos(), pdfs_generados_destino)
    movidos_respaldo = mover_archivos(carpeta_backup_facturas(), pdfs_respaldo_destino)

    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM orden_detalles")
    cur.execute("DELETE FROM ordenes")
    cur.execute("DELETE FROM cobros")
    cur.execute("DELETE FROM documentos")
    cur.execute("DELETE FROM sqlite_sequence WHERE name IN ('orden_detalles', 'ordenes', 'cobros', 'documentos')")
    conn.commit()
    conn.close()

    borrados_nube = 0
    if usar_backend_nube():
        resultado_nube = api_delete_json("/documentos") or {}
        borrados_nube = int(resultado_nube.get("borrados") or 0)

    return {
        "carpeta": carpeta_destino,
        "documentos_generados": movidos_generados,
        "respaldos_pdf": movidos_respaldo,
        "borrados_nube": borrados_nube,
        "respaldo_nube": bool(respaldo_nube),
    }


def abrir_archivo_generado(ruta):
    if not ruta or not os.path.exists(ruta):
        return False

    try:
        if sys.platform.startswith("darwin"):
            try:
                subprocess.Popen(["open", "-a", "Preview", ruta])
            except Exception:
                subprocess.Popen(["open", ruta])
        elif os.name == "nt":
            os.startfile(ruta)
        else:
            subprocess.Popen(["xdg-open", ruta])
        return True
    except Exception:
        return False


def imprimir_archivo_generado(ruta):
    if not ruta or not os.path.exists(ruta):
        return False, "El archivo no existe."

    try:
        resultado = subprocess.run(["lp", ruta], capture_output=True, text=True)
        if resultado.returncode == 0:
            return True, (resultado.stdout or "Documento enviado a la impresora.").strip()
        return False, (resultado.stderr or "No se pudo enviar a imprimir.").strip()
    except Exception as exc:
        return False, str(exc)


def monto(valor):
    return f"RD$ {float(valor):,.2f}"


def resolver_descuento(subtotal, descuento_valor):
    descuento_valor = max(0.0, float(descuento_valor or 0))
    if descuento_valor <= 100:
        descuento_pct = descuento_valor
        descuento_monto = subtotal * (descuento_pct / 100.0)
    else:
        descuento_monto = min(subtotal, descuento_valor)
        descuento_pct = (descuento_monto / subtotal * 100.0) if subtotal else 0.0
    return descuento_pct, descuento_monto


def calcular_resumen_financiero(subtotal, descuento_valor=0, abono=0):
    subtotal = max(0.0, float(subtotal or 0))
    descuento_pct, descuento_monto = resolver_descuento(subtotal, descuento_valor)
    base_imponible = max(0.0, subtotal - descuento_monto)
    itbis = base_imponible * 0.18
    total_final = base_imponible + itbis
    abono = max(0.0, float(abono or 0))
    restante = total_final - abono
    return {
        "subtotal": subtotal,
        "descuento_pct": descuento_pct,
        "descuento_monto": descuento_monto,
        "itbis": itbis,
        "total_final": total_final,
        "abono": abono,
        "restante": restante,
    }


def normalizar_metodo_pago(texto):
    texto = (texto or "").strip().lower()
    if texto in ("tarjeta", "targeta"):
        return "Tarjeta"
    if texto == "efectivo":
        return "Efectivo"
    if texto == "transferencia":
        return "Transferencia"
    return "Pendiente"


def dibujar_logo_pdf(c, x, y):
    if os.path.exists(LOGO_NEGOCIO):
        try:
            logo = ImageReader(LOGO_NEGOCIO)
            c.drawImage(logo, x, y - 92, width=120, height=92, preserveAspectRatio=True, mask="auto")
            return
        except Exception:
            pass

    color_principal = colors.HexColor("#57C7C3")
    color_secundario = colors.HexColor("#2E6A56")

    c.saveState()
    c.setLineWidth(1.5)
    for i, tam in enumerate([34, 26, 18, 10]):
        c.setStrokeColor(color_principal if i < 2 else color_secundario)
        c.rect(x + i * 4, y - tam - i * 4, tam, tam)
    c.setFont("Helvetica-Bold", 7)
    c.setFillColor(color_secundario)
    c.drawString(x, y - 44, "ENMARCADOS PF")
    c.restoreState()


def dibujar_logo_encajado(c, x, y, w, h):
    borde = colors.HexColor("#D7DED4")
    c.saveState()
    c.setFillColor(colors.white)
    c.setStrokeColor(borde)
    c.roundRect(x, y, w, h, 10, fill=1, stroke=1)

    if os.path.exists(LOGO_NEGOCIO):
        try:
            logo = ImageReader(LOGO_NEGOCIO)
            img_w, img_h = logo.getSize()
            escala = max(w / img_w, h / img_h)
            draw_w = img_w * escala
            draw_h = img_h * escala
            draw_x = x + (w - draw_w) / 2
            draw_y = y + (h - draw_h) / 2

            path = c.beginPath()
            path.roundRect(x, y, w, h, 10)
            c.clipPath(path, stroke=0, fill=0)
            c.drawImage(logo, draw_x, draw_y, width=draw_w, height=draw_h, preserveAspectRatio=True, mask="auto")
            c.restoreState()
            return
        except Exception:
            c.restoreState()
            dibujar_logo_pdf(c, x + 16, y + h - 12)
            return

    c.restoreState()
    dibujar_logo_pdf(c, x + 16, y + h - 12)


def dibujar_bloque_factura(c, x, y, w, titulo, lineas, color_barra):
    alto = 20 + max(44, 14 * max(1, len(lineas)))
    c.saveState()
    c.setStrokeColor(color_barra)
    c.setLineWidth(1)
    c.rect(x, y - alto, w, alto, fill=0, stroke=1)
    c.setFillColor(color_barra)
    c.rect(x, y - 18, w, 18, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x + 6, y - 12, titulo)
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 8.5)
    cursor_y = y - 32
    for linea in lineas:
        c.drawString(x + 6, cursor_y, linea[:56])
        cursor_y -= 12
    c.restoreState()
    return y - alto


class SelectorFechaPopup(ctk.CTkToplevel):
    def __init__(self, master, callback, fecha_inicial=""):
        super().__init__(master)
        self.callback = callback
        self.title("Seleccionar fecha")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()

        try:
            fecha = datetime.strptime(fecha_inicial, "%Y-%m-%d")
            self.anio = fecha.year
            self.mes = fecha.month
        except Exception:
            self.anio, self.mes = fecha_mes_actual()

        self.header = ctk.CTkFrame(self, fg_color="transparent")
        self.header.pack(fill="x", padx=12, pady=(12, 4))

        ctk.CTkButton(self.header, text="<", width=34, command=self.mes_anterior).pack(side="left")
        self.titulo_label = ctk.CTkLabel(self.header, text="", font=ctk.CTkFont(size=16, weight="bold"))
        self.titulo_label.pack(side="left", expand=True)
        ctk.CTkButton(self.header, text=">", width=34, command=self.mes_siguiente).pack(side="right")

        self.grid_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.grid_frame.pack(padx=12, pady=(0, 12))
        self.renderizar()

    def mes_anterior(self):
        self.mes -= 1
        if self.mes < 1:
            self.mes = 12
            self.anio -= 1
        self.renderizar()

    def mes_siguiente(self):
        self.mes += 1
        if self.mes > 12:
            self.mes = 1
            self.anio += 1
        self.renderizar()

    def seleccionar(self, dia):
        fecha = f"{self.anio:04d}-{self.mes:02d}-{dia:02d}"
        self.callback(fecha)
        self.destroy()

    def renderizar(self):
        for child in self.grid_frame.winfo_children():
            child.destroy()

        nombre_mes = pycalendar.month_name[self.mes].capitalize()
        self.titulo_label.configure(text=f"{nombre_mes} {self.anio}")

        dias = ["Lu", "Ma", "Mi", "Ju", "Vi", "Sa", "Do"]
        for col, nombre in enumerate(dias):
            ctk.CTkLabel(self.grid_frame, text=nombre, width=38).grid(row=0, column=col, padx=2, pady=2)

        cal = pycalendar.Calendar(firstweekday=0)
        semanas = cal.monthdayscalendar(self.anio, self.mes)
        for row_idx, semana in enumerate(semanas, start=1):
            for col_idx, dia in enumerate(semana):
                if dia == 0:
                    ctk.CTkLabel(self.grid_frame, text="", width=38, height=30).grid(row=row_idx, column=col_idx, padx=2, pady=2)
                else:
                    ctk.CTkButton(
                        self.grid_frame,
                        text=str(dia),
                        width=38,
                        height=30,
                        fg_color="white",
                        text_color="black",
                        border_width=1,
                        border_color=COLOR_BORDE,
                        hover_color=COLOR_PANEL,
                        command=lambda d=dia: self.seleccionar(d),
                    ).grid(row=row_idx, column=col_idx, padx=2, pady=2)


def dibujar_encabezado_documento(c, titulo, numero_doc, fecha, cliente, telefono, rnc, direccion):
    color_principal = colors.HexColor("#57C7C3")
    color_oscuro = colors.HexColor("#1D4E3F")
    width, height = letter

    c.setStrokeColor(color_principal)
    c.setLineWidth(3)
    c.rect(18, 18, width - 36, height - 36)

    dibujar_logo_pdf(c, 28, height - 28)

    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 24)
    c.drawCentredString(width / 2, height - 40, "ENMARCADOS")
    c.setFillColor(color_principal)
    c.drawString(width / 2 + 110, height - 40, "PF")

    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 15)
    c.drawCentredString(width / 2, height - 70, titulo)

    y = height - 115
    c.setLineWidth(1)
    c.setStrokeColor(colors.black)
    campos = [
        ("CLIENTE", cliente or "N/D"),
        ("DIRECCIÓN", direccion or "N/D"),
        ("TELÉFONO", telefono or "N/D"),
        ("RNC", rnc or "N/D"),
        ("FECHA", fecha),
        ("NÚMERO", str(numero_doc)),
    ]
    for etiqueta, valor in campos:
        c.rect(60, y - 12, 95, 16)
        c.rect(155, y - 12, 245, 16)
        c.setFont("Helvetica-Bold", 7)
        c.drawString(64, y - 7, etiqueta)
        c.setFont("Helvetica", 9)
        c.drawString(160, y - 7, str(valor)[:45])
        y -= 16

    return y - 10


def dibujar_tarjeta_resumen(c, x, y, w, h, titulo, valor, prefijo=""):
    c.saveState()
    c.setFillColor(colors.HexColor("#F7FBFB"))
    c.setStrokeColor(colors.HexColor("#D7E4E7"))
    c.roundRect(x, y, w, h, 12, stroke=1, fill=1)
    c.setFillColor(colors.HexColor("#5E6B73"))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(x + 12, y + h - 16, titulo)
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(x + 12, y + 12, f"{prefijo}{valor}")
    c.restoreState()


def dibujar_header_factura_moderno(c, titulo, numero_doc, fecha, cliente, telefono, rnc, direccion):
    width, height = letter
    c.setFillColor(colors.white)
    c.rect(0, 0, width, height, fill=1, stroke=0)

    c.setStrokeColor(colors.HexColor(COLOR_PRINCIPAL))
    c.setLineWidth(6)
    c.line(30, height - 40, width - 30, height - 40)

    dibujar_logo_pdf(c, 34, height - 24)

    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 28)
    c.drawString(38, height - 126, titulo)

    c.setFillColor(colors.HexColor(COLOR_TEXTO_SUAVE))
    c.setFont("Helvetica", 11)
    c.drawString(38, height - 142, f"Numero: INV-{numero_doc:06d}")
    c.drawString(38, height - 157, f"Fecha: {fecha}")

    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(colors.black)
    c.drawRightString(width - 40, height - 64, "ENMARCADOS PF")
    c.setFont("Helvetica", 10)
    c.setFillColor(colors.HexColor(COLOR_TEXTO_SUAVE))
    c.drawRightString(width - 40, height - 80, "Tu tienda de enmarcados")
    c.drawRightString(width - 40, height - 96, f"Telefono: {telefono or 'N/D'}")
    c.drawRightString(width - 40, height - 112, f"RNC: {rnc or 'N/D'}")

    c.setFillColor(colors.HexColor("#F7FBFB"))
    c.setStrokeColor(colors.HexColor("#D7E4E7"))
    c.roundRect(34, height - 245, 255, 74, 12, stroke=1, fill=1)
    c.setFillColor(colors.HexColor(COLOR_TEXTO_SUAVE))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(46, height - 187, "Cliente")
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(46, height - 204, cliente or "N/D")
    c.setFont("Helvetica", 9)
    c.drawString(46, height - 220, direccion or "Direccion no registrada")

    c.setFillColor(colors.HexColor("#F7FBFB"))
    c.roundRect(width - 220, height - 245, 186, 74, 12, stroke=1, fill=1)
    c.setFillColor(colors.HexColor(COLOR_TEXTO_SUAVE))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(width - 208, height - 187, "Detalle")
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 10)
    c.drawString(width - 208, height - 204, f"Documento: {titulo}")
    c.drawString(width - 208, height - 220, f"Contacto: {telefono or 'N/D'}")

    return height - 280


def obtener_detalles_de_orden(orden_id):
    if usar_backend_nube():
        docs = api_get_json("/documentos", params={"limit": 300}) or []
        for doc in docs:
            for orden in doc.get("ordenes", []):
                if orden.get("id") == orden_id:
                    return [
                        (
                            item.get("cantidad", 0),
                            item.get("codigo_material"),
                            item.get("descripcion_material"),
                            item.get("ancho", 0),
                            item.get("largo", 0),
                            item.get("pies", 0),
                            item.get("precio", 0),
                            item.get("subtotal", 0),
                            item.get("total", 0),
                        )
                        for item in orden.get("detalles", [])
                    ]
        return []
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT cantidad, codigo_material, descripcion_material,
               ancho, largo, pies, precio, subtotal, total
        FROM orden_detalles
        WHERE orden_id = ?
        ORDER BY id
    """, (orden_id,))
    filas = cur.fetchall()
    conn.close()
    return filas


def obtener_ordenes_completas_de_documento(documento_id):
    if usar_backend_nube():
        data = api_get_json(f"/documentos/{documento_id}")
        if not data or data.get("error"):
            return []
        ordenes = []
        for orden in data.get("ordenes", []):
            ordenes.append({
                "id": orden.get("id"),
                "a_enmarcar": orden.get("a_enmarcar"),
                "notas": orden.get("notas") or "",
                "ancho": orden.get("ancho", 0),
                "largo": orden.get("largo", 0),
                "total_orden": orden.get("total_orden", 0),
                "detalles": [
                    (
                        item.get("cantidad", 0),
                        item.get("codigo_material"),
                        item.get("descripcion_material"),
                        item.get("ancho", 0),
                        item.get("largo", 0),
                        item.get("pies", 0),
                        item.get("precio", 0),
                        item.get("subtotal", 0),
                        item.get("total", 0),
                    )
                    for item in orden.get("detalles", [])
                ],
            })
        return ordenes

    ordenes = []

    for orden_id, a_enmarcar, notas, ancho, largo, total_orden in obtener_ordenes_de_documento(documento_id):
        ordenes.append({
            "id": orden_id,
            "a_enmarcar": a_enmarcar,
            "notas": notas or "",
            "ancho": ancho,
            "largo": largo,
            "total_orden": total_orden,
            "detalles": obtener_detalles_de_orden(orden_id),
        })

    return ordenes


def obtener_cantidad_principal_orden(orden):
    detalles = orden.get("detalles") or []
    if detalles:
        try:
            return float(detalles[0][0])
        except (TypeError, ValueError, IndexError):
            pass
    return float(orden.get("cantidad", 1) or 1)


def generar_pdf_documento(documento_id):
    documento = obtener_documento(documento_id)
    if not documento:
        return None

    _, tipo, numero_doc, _, nombre, telefono, rnc, direccion, fecha, subtotal, descuento, itbis, total_final, _, metodo_pago, _retirado, _fecha_entrega = documento
    ordenes = obtener_ordenes_completas_de_documento(documento_id)
    total_cobrado = obtener_total_cobrado_documento(documento_id)
    resumen = calcular_resumen_financiero(subtotal, descuento, total_cobrado)

    nombre_base = f"{fecha}_{tipo.upper()}_{numero_doc:06d}_{limpiar_nombre_archivo(nombre)}"
    ruta = os.path.join(carpeta_documentos(), f"{nombre_base}.pdf")

    c = canvas.Canvas(ruta, pagesize=letter)
    width, height = letter
    color_barra = colors.HexColor(COLOR_PRINCIPAL)
    color_linea = colors.HexColor("#CFE7E7")
    color_fila = colors.HexColor("#F4FBFB")
    color_pagado = colors.HexColor(COLOR_PRINCIPAL)
    color_pendiente = colors.HexColor("#D84949")
    titulo = "FACTURA" if tipo == "factura" else "COTIZACION"

    def dibujar_pie():
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(width / 2, 46, f"Instagram: {NEGOCIO_INSTAGRAM}")
        c.setFont("Helvetica", 10)
        c.drawCentredString(width / 2, 32, NEGOCIO_DIRECCION)
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(width / 2, 18, f"TELEFONO: {NEGOCIO_TELEFONO}")
        c.setFont("Helvetica", 7.5)
        c.drawCentredString(width / 2, 8, "**No nos hacemos responsables de trabajos no retirados despues de 30 dias")

    def dibujar_plantilla():
        c.setFillColor(colors.white)
        c.rect(0, 0, width, height, fill=1, stroke=0)

        dibujar_logo_encajado(c, 36, height - 126, 136, 88)

        c.setFillColor(colors.black)
        c.setFont("Times-Bold", 32)
        c.drawCentredString(width / 2, height - 54, titulo)

        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 17)
        c.drawString(36, height - 154, "Enmarcados PF")
        c.setFont("Helvetica", 9)
        c.drawString(36, height - 170, NEGOCIO_DIRECCION)
        c.drawString(36, height - 184, f"Telefono: {NEGOCIO_TELEFONO}")
        c.drawString(36, height - 198, f"RNC: {NEGOCIO_RNC}")
        c.drawString(36, height - 212, f"Instagram: {NEGOCIO_INSTAGRAM}")

        info_x = width - 228
        info_y = height - 76
        c.setStrokeColor(colors.black)
        c.setLineWidth(1)
        c.rect(info_x, info_y - 50, 192, 50, fill=0, stroke=1)
        filas_info = [
            ("ID FACTURA", f"{numero_doc:06d}"),
            ("FECHA", fecha),
            ("ID", str(documento_id)),
        ]
        fila_h = 16.66
        for i, (label, valor) in enumerate(filas_info):
            top_y = info_y - i * fila_h
            c.setFillColor(color_barra)
            c.rect(info_x, top_y - fila_h, 84, fila_h, fill=1, stroke=1)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 8)
            c.drawCentredString(info_x + 42, top_y - 11, label)
            c.setFillColor(colors.white)
            c.rect(info_x + 84, top_y - fila_h, 108, fila_h, fill=1, stroke=1)
            c.setFillColor(colors.black)
            c.setFont("Helvetica", 8)
            c.drawCentredString(info_x + 138, top_y - 11, str(valor)[:24])

        y_bloques = height - 226
        dibujar_bloque_factura(
            c, 36, y_bloques, 520, "DATOS DEL CLIENTE",
            [
                f"Nombre: {nombre or 'Cliente sin nombre'}",
                f"Direccion: {direccion or 'Direccion no registrada'}",
                f"Telefono: {telefono or '-'}",
                f"RNC: {rnc or '-'}",
            ],
            color_barra
        )

        y_tabla = height - 320
        columnas = [
            (36, 90, "CANTIDAD"),
            (126, 330, "DESCRIPCION"),
            (456, 120, "TOTAL"),
        ]
        for x, w, txt in columnas:
            c.setFillColor(color_barra)
            c.rect(x, y_tabla - 18, w, 18, fill=1, stroke=1)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 8.2)
            c.drawCentredString(x + w / 2, y_tabla - 12, txt)

        return y_tabla - 18

    y = dibujar_plantilla()

    for idx, orden in enumerate(ordenes, start=1):
        if y < 156:
            dibujar_pie()
            c.showPage()
            y = dibujar_plantilla()

        cantidad_general = obtener_cantidad_principal_orden(orden)
        cantidad_txt = str(int(cantidad_general) if cantidad_general == int(cantidad_general) else round(cantidad_general, 2))
        unitario = float(orden["total_orden"]) / cantidad_general if cantidad_general else float(orden["total_orden"])

        c.setFillColor(color_fila if idx % 2 == 0 else colors.white)
        c.setStrokeColor(color_linea)
        c.rect(36, y - 18, 540, 18, fill=1, stroke=1)
        c.setFillColor(colors.black)
        c.setFont("Helvetica", 8.2)
        c.drawCentredString(81, y - 12, cantidad_txt)
        c.drawCentredString(291, y - 12, str(orden['a_enmarcar'])[:52])
        c.drawRightString(568, y - 12, f"{float(orden['total_orden']):,.2f}")
        y -= 18

    while y > 164:
        c.setFillColor(color_fila if int((height - y) / 18) % 2 == 0 else colors.white)
        c.setStrokeColor(color_linea)
        c.rect(36, y - 18, 540, 18, fill=1, stroke=1)
        y -= 18

    notas_y = 142
    c.setStrokeColor(colors.black)
    c.rect(36, notas_y - 62, 316, 62, fill=0, stroke=1)
    c.setFillColor(color_barra)
    c.rect(36, notas_y - 14, 316, 14, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 7.5)
    c.drawCentredString(194, notas_y - 10, "NOTAS DE PAGO")
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 7)
    c.drawString(42, notas_y - 28, "Para cualquier consulta o inconveniente, por favor contactanos.")
    c.drawString(42, notas_y - 40, "Metodo de pago actual: " + normalizar_metodo_pago(metodo_pago))
    c.drawString(42, notas_y - 52, "Gracias por confiar en Enmarcados PF.")

    resumen_x = 386
    resumen_y = 142
    resumen_filas = [
        ("SUBTOTAL", resumen["subtotal"]),
        ("DESCUENTO", resumen["descuento_monto"]),
        ("ITBIS", resumen["itbis"]),
        ("TOTAL", resumen["total_final"]),
        ("ABONO", resumen["abono"]),
    ]
    c.setStrokeColor(colors.black)
    c.rect(resumen_x, resumen_y - 74, 190, 74, fill=0, stroke=1)
    for i, (label, valor) in enumerate(resumen_filas):
        fila_y = resumen_y - i * 14.8
        c.line(resumen_x, fila_y - 14.8, resumen_x + 190, fila_y - 14.8)
        c.setFont("Helvetica-Bold", 7.4)
        c.drawString(resumen_x + 8, fila_y - 10, label)
        c.setFont("Helvetica", 7.4)
        c.drawRightString(resumen_x + 182, fila_y - 10, f"{float(valor):,.2f}")

    balance_color = color_pagado if resumen["restante"] <= 0 and resumen["total_final"] > 0 else color_pendiente
    c.setFillColor(balance_color)
    c.rect(resumen_x, 48, 190, 18, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.rect(resumen_x, 48, 190, 18, fill=0, stroke=1)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(resumen_x + 8, 54, "BALANCE PENDIENTE")
    c.drawRightString(resumen_x + 182, 54, f"{max(0, resumen['restante']):,.2f}")

    if resumen["restante"] <= 0 and resumen["total_final"] > 0:
        c.saveState()
        c.translate(468, 108)
        c.rotate(12)
        c.setStrokeColor(color_pagado)
        c.setFillColor(color_pagado)
        c.setLineWidth(2)
        c.roundRect(-54, -14, 108, 28, 8, stroke=1, fill=0)
        c.setFont("Helvetica-Bold", 15)
        c.drawCentredString(0, -4, "PAGADA")
        c.restoreState()
        c.setFillColor(color_pagado)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(36, 74, "PAGO")
    else:
        c.setFillColor(color_pendiente)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(36, 74, "PENDIENTE DE PAGO")

    dibujar_pie()

    c.save()
    return ruta


def generar_pdf_ordenes(documento_id):
    documento = obtener_documento(documento_id)
    if not documento:
        return None

    _, tipo, numero_doc, _, nombre, telefono, rnc, direccion, fecha, subtotal, descuento, _, _, _, _metodo_pago, _retirado, fecha_entrega = documento
    ordenes = obtener_ordenes_completas_de_documento(documento_id)
    resumen = calcular_resumen_financiero(subtotal, descuento)

    nombre_base = f"{fecha}_ORDENES_{tipo.upper()}_{numero_doc:06d}_{limpiar_nombre_archivo(nombre)}"
    ruta = os.path.join(carpeta_documentos(), f"{nombre_base}.pdf")

    c = canvas.Canvas(ruta, pagesize=letter)
    width, height = letter
    color_principal = colors.HexColor(COLOR_PRINCIPAL)
    color_panel = colors.HexColor("#F7FBFB")
    color_linea = colors.HexColor("#D7E4E7")

    columnas = [
        (34, 22, "N"),
        (56, 32, "CANT"),
        (88, 46, "COD"),
        (134, 154, "DESCRIPCION"),
        (288, 118, "A ENMARCAR"),
        (406, 32, "AN"),
        (438, 32, "LA"),
        (470, 40, "P2"),
        (510, 58, "TOTAL"),
    ]
    row_h = 12

    def dibujar_footer():
        c.setStrokeColor(colors.HexColor("#1F2B3A"))
        c.line(34, 34, width - 34, 34)
        c.setFillColor(colors.HexColor("#1F2B3A"))
        c.setFont("Helvetica-Bold", 8)
        c.drawString(34, 22, "Enmarcados PF")
        c.setFont("Helvetica", 7)
        c.drawRightString(width - 34, 22, "Instagram: @enmarcadospfrd | Tel. 809 569 2000")

    def nueva_pagina():
        c.setFillColor(colors.white)
        c.rect(0, 0, width, height, fill=1, stroke=0)
        y_local = dibujar_encabezado_documento(
            c, "ORDEN DE TRABAJO", numero_doc, fecha, nombre, telefono, rnc, direccion
        )
        c.setFillColor(color_principal)
        c.roundRect(34, y_local - 18, width - 68, 16, 8, fill=1, stroke=0)
        for x, ancho_col, titulo_col in columnas:
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 6.2)
            c.drawCentredString(x + ancho_col / 2, y_local - 12, titulo_col)
        c.setFillColor(colors.black)
        dibujar_footer()
        return y_local - 24

    y = nueva_pagina()

    for indice_orden, orden in enumerate(ordenes, start=1):
        lineas_necesarias = len(orden["detalles"]) + 5
        espacio_necesario = lineas_necesarias * row_h + 34
        if y - espacio_necesario < 52:
            c.showPage()
            y = nueva_pagina()

        c.setFillColor(color_panel)
        c.roundRect(34, y - 20, width - 68, 18, 8, fill=1, stroke=0)
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(40, y - 13, f"Orden {indice_orden}")
        c.setFont("Helvetica", 7.2)
        c.drawString(88, y - 13, str(orden["a_enmarcar"])[:70])
        c.drawRightString(width - 40, y - 13, f'{orden["ancho"]} x {orden["largo"]}')
        y -= 22

        if orden.get("notas"):
            c.setFillColor(colors.HexColor(COLOR_TEXTO_SUAVE))
            c.setFont("Helvetica-Oblique", 6.5)
            c.drawString(40, y - 6, f"Notas: {str(orden['notas'])[:92]}")
            y -= 12
            c.setFillColor(colors.black)

        for indice_detalle, detalle in enumerate(orden["detalles"], start=1):
            cantidad, codigo_material, descripcion_material, ancho_d, largo_d, pies, _precio, _subtotal_d, total = detalle

            if y - row_h < 52:
                c.showPage()
                y = nueva_pagina()

            c.setFillColor(colors.white if indice_detalle % 2 else color_panel)
            c.roundRect(34, y - row_h, width - 68, row_h - 1, 4, fill=1, stroke=0)
            c.setFillColor(colors.black)
            c.setFont("Helvetica", 6.3)
            valores = [
                str(indice_detalle),
                str(int(cantidad) if cantidad == int(cantidad) else round(cantidad, 2)),
                str(codigo_material),
                str(descripcion_material)[:28],
                str(orden["a_enmarcar"])[:22],
                str(ancho_d),
                str(largo_d),
                f"{pies:.2f}",
                f"{float(total):,.2f}",
            ]
            for (x, ancho_col, _), texto in zip(columnas, valores):
                if x == 510:
                    c.drawRightString(x + ancho_col - 4, y - 8, texto)
                elif x in (34, 56, 88, 406, 438, 470):
                    c.drawCentredString(x + ancho_col / 2, y - 8, texto)
                else:
                    c.drawString(x + 3, y - 8, texto)
            y -= row_h

        y -= 4
        c.setStrokeColor(color_linea)
        c.line(330, y, width - 40, y)
        y -= 10
        c.setFont("Helvetica-Bold", 7)
        c.drawString(374, y, "Total orden")
        c.drawRightString(width - 40, y, monto(orden["total_orden"]))
        y -= 12

    if ordenes:
        c.setFillColor(color_panel)
        c.roundRect(330, 48, width - 364, 34, 8, fill=1, stroke=0)
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 8)
        c.drawString(340, 69, "Resumen documento")
        c.setFont("Helvetica", 7)
        c.drawString(340, 57, f"Subtotal: {monto(resumen['subtotal'])}")
        c.drawString(430, 57, f"Descuento: {resumen['descuento_pct']:.1f}%")
        c.drawRightString(width - 40, 57, f"Total: {monto(resumen['total_final'])}")

    c.save()
    return ruta


def exportar_documentos(documento_id):
    descuento = 0.0
    documento = obtener_documento(documento_id)
    if documento:
        descuento = float(documento[10] or 0)
        actualizar_totales_documento(documento_id, descuento)

    ruta_factura = generar_pdf_documento(documento_id)
    ruta_orden = generar_pdf_ordenes(documento_id)
    guardar_copia_backup(ruta_factura)
    guardar_copia_backup(ruta_orden)
    return ruta_factura, ruta_orden

# =========================
# BASE DE DATOS
# =========================
def conectar_db():
    return sqlite3.connect(DB_NAME)


def columna_existe(cur, tabla, columna):
    cur.execute(f"PRAGMA table_info({tabla})")
    columnas = [fila[1] for fila in cur.fetchall()]
    return columna in columnas


def crear_tablas():
    if getattr(sys, "frozen", False) and not os.path.exists(DB_NAME):
        origen_db = ruta_recurso("facturacion.db")
        if os.path.exists(origen_db):
            os.makedirs(os.path.dirname(DB_NAME), exist_ok=True)
            shutil.copy2(origen_db, DB_NAME)

    conn = conectar_db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS tarifas (
            codigo INTEGER PRIMARY KEY,
            nombre TEXT NOT NULL,
            precio REAL NOT NULL,
            extra REAL NOT NULL DEFAULT 0
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS clientes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL UNIQUE,
            telefono TEXT,
            rnc TEXT,
            direccion TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS documentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo TEXT NOT NULL,
            numero_doc INTEGER,
            cliente_id INTEGER NOT NULL,
            fecha TEXT NOT NULL,
            fecha_entrega TEXT NOT NULL DEFAULT '',
            subtotal REAL NOT NULL DEFAULT 0,
            descuento REAL NOT NULL DEFAULT 0,
            itbis REAL NOT NULL DEFAULT 0,
            total_final REAL NOT NULL DEFAULT 0,
            cerrado INTEGER NOT NULL DEFAULT 0,
            metodo_pago TEXT NOT NULL DEFAULT 'Pendiente',
            retirado INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY (cliente_id) REFERENCES clientes(id)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS ordenes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            factura_id INTEGER,
            documento_id INTEGER NOT NULL,
            a_enmarcar TEXT NOT NULL,
            notas TEXT NOT NULL DEFAULT '',
            ancho REAL NOT NULL,
            largo REAL NOT NULL,
            total_orden REAL NOT NULL DEFAULT 0,
            FOREIGN KEY (documento_id) REFERENCES documentos(id)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS orden_detalles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            orden_id INTEGER NOT NULL,
            cantidad REAL NOT NULL,
            codigo_material INTEGER NOT NULL,
            descripcion_material TEXT NOT NULL,
            ancho REAL NOT NULL,
            largo REAL NOT NULL,
            pies REAL NOT NULL,
            precio REAL NOT NULL,
            subtotal REAL NOT NULL,
            total REAL NOT NULL,
            FOREIGN KEY (orden_id) REFERENCES ordenes(id)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS cobros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            documento_id INTEGER NOT NULL,
            numero_doc INTEGER NOT NULL,
            cliente_nombre TEXT NOT NULL,
            fecha TEXT NOT NULL,
            monto REAL NOT NULL DEFAULT 0,
            metodo_pago TEXT NOT NULL DEFAULT 'Pendiente',
            referencia TEXT,
            pagado_total INTEGER NOT NULL DEFAULT 0,
            FOREIGN KEY (documento_id) REFERENCES documentos(id)
        )
    """)

    # migraciones simples
    if not columna_existe(cur, "documentos", "numero_doc"):
        cur.execute("ALTER TABLE documentos ADD COLUMN numero_doc INTEGER")
    if not columna_existe(cur, "documentos", "cerrado"):
        cur.execute("ALTER TABLE documentos ADD COLUMN cerrado INTEGER NOT NULL DEFAULT 0")
    if not columna_existe(cur, "documentos", "metodo_pago"):
        cur.execute("ALTER TABLE documentos ADD COLUMN metodo_pago TEXT NOT NULL DEFAULT 'Pendiente'")
    if not columna_existe(cur, "documentos", "retirado"):
        cur.execute("ALTER TABLE documentos ADD COLUMN retirado INTEGER NOT NULL DEFAULT 0")
    if not columna_existe(cur, "documentos", "fecha_entrega"):
        cur.execute("ALTER TABLE documentos ADD COLUMN fecha_entrega TEXT NOT NULL DEFAULT ''")
    if columna_existe(cur, "ordenes", "factura_id") and not columna_existe(cur, "ordenes", "documento_id"):
        cur.execute("ALTER TABLE ordenes ADD COLUMN documento_id INTEGER")
        cur.execute("""
            UPDATE ordenes
            SET documento_id = factura_id
            WHERE documento_id IS NULL
        """)
    if columna_existe(cur, "ordenes", "documento_id") and not columna_existe(cur, "ordenes", "factura_id"):
        cur.execute("ALTER TABLE ordenes ADD COLUMN factura_id INTEGER")
        cur.execute("""
            UPDATE ordenes
            SET factura_id = documento_id
            WHERE factura_id IS NULL
        """)
    if not columna_existe(cur, "ordenes", "notas"):
        cur.execute("ALTER TABLE ordenes ADD COLUMN notas TEXT NOT NULL DEFAULT ''")

    conn.commit()
    conn.close()


# =========================
# CLIENTES
# =========================
def guardar_cliente_db(nombre, telefono, rnc, direccion):
    if usar_backend_nube():
        data = api_post_json("/clientes", {
            "nombre": nombre,
            "telefono": telefono,
            "rnc": rnc,
            "direccion": direccion,
        })
        return data.get("id") if isinstance(data, dict) else None
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO clientes (nombre, telefono, rnc, direccion)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(nombre) DO UPDATE SET
            telefono=excluded.telefono,
            rnc=excluded.rnc,
            direccion=excluded.direccion
    """, (nombre, telefono, rnc, direccion))
    conn.commit()
    conn.close()


def obtener_cliente_por_nombre(nombre):
    if usar_backend_nube():
        try:
            data = api_get_json(f"/clientes/by-name/{quote(nombre)}")
            if data and not data.get("error"):
                return (
                    data.get("id"),
                    data.get("nombre"),
                    data.get("telefono"),
                    data.get("rnc"),
                    data.get("direccion"),
                )
        except Exception:
            pass
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, nombre, telefono, rnc, direccion
        FROM clientes
        WHERE nombre = ?
    """, (nombre,))
    fila = cur.fetchone()
    conn.close()
    return fila


def obtener_todos_clientes():
    if usar_backend_nube():
        try:
            data = api_get_json("/clientes")
            if isinstance(data, list):
                return [
                    (
                        item.get("id"),
                        item.get("nombre"),
                        item.get("telefono"),
                        item.get("rnc"),
                        item.get("direccion"),
                    )
                    for item in data
                ]
        except Exception:
            pass
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, nombre, telefono, rnc, direccion
        FROM clientes
        ORDER BY nombre
    """)
    filas = cur.fetchall()
    conn.close()
    return filas

def eliminar_cliente_por_nombre(nombre):
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM clientes WHERE nombre = ?", (nombre,))
    cambios = cur.rowcount
    conn.commit()
    conn.close()
    return cambios


def borrar_todos_los_clientes():
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM clientes")
    conn.commit()
    conn.close()


def importar_clientes_desde_excel(ruta="clientes.xlsx"):
    from openpyxl import load_workbook
    import os

    if not os.path.exists(ruta):
        return 0, 0, ["No se encontró el archivo clientes.xlsx"]

    wb = load_workbook(ruta)
    hoja = wb.active

    insertados = 0
    saltados = 0
    errores = []

    for i, fila in enumerate(hoja.iter_rows(min_row=2), start=2):
        try:
            nombre = fila[0].value
            telefono = fila[1].value
            rnc = fila[2].value
            direccion = fila[3].value

            if not nombre:
                saltados += 1
                continue

            nombre = str(nombre).strip()
            telefono = "" if telefono is None else str(telefono).strip()
            rnc = "" if rnc is None else str(rnc).strip()
            direccion = "" if direccion is None else str(direccion).strip()

            guardar_cliente_db(nombre, telefono, rnc, direccion)
            insertados += 1

        except Exception as e:
            errores.append(f"Fila {i}: {str(e)}")
            saltados += 1

    return insertados, saltados, errores

def importar_clientes_desde_excel(ruta="clientes.xlsx"):
    from openpyxl import load_workbook
    import os

    if not os.path.exists(ruta):
        return 0, 0, ["No se encontró el archivo clientes.xlsx"]

    wb = load_workbook(ruta)
    hoja = wb.active

    insertados = 0
    saltados = 0
    errores = []

    for i, fila in enumerate(hoja.iter_rows(min_row=2), start=2):
        try:
            nombre = fila[0].value
            telefono = fila[1].value
            rnc = fila[2].value
            direccion = fila[3].value

            if not nombre:
                saltados += 1
                continue

            nombre = str(nombre).strip()
            telefono = "" if telefono is None else str(telefono).strip()
            rnc = "" if rnc is None else str(rnc).strip()
            direccion = "" if direccion is None else str(direccion).strip()

            guardar_cliente_db(nombre, telefono, rnc, direccion)
            insertados += 1

        except Exception as e:
            errores.append(f"Fila {i}: {str(e)}")
            saltados += 1

    return insertados, saltados, errores

def eliminar_cliente_por_nombre(nombre):
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM clientes WHERE nombre = ?", (nombre,))
    cambios = cur.rowcount
    conn.commit()
    conn.close()
    return cambios


def borrar_todos_los_clientes():
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM clientes")
    conn.commit()
    conn.close()

# =========================
# TARIFAS
# =========================
def guardar_tarifa_db(codigo, nombre, precio, extra):
    if usar_backend_nube():
        api_post_json("/tarifas", {
            "codigo": int(codigo),
            "nombre": nombre,
            "precio": float(precio),
            "extra": float(extra),
        })
        return
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO tarifas (codigo, nombre, precio, extra)
        VALUES (?, ?, ?, ?)
        ON CONFLICT(codigo) DO UPDATE SET
            nombre=excluded.nombre,
            precio=excluded.precio,
            extra=excluded.extra
    """, (codigo, nombre, precio, extra))
    conn.commit()
    conn.close()
def limpiar_numero(texto):
    texto = str(texto).strip()
    if not texto:
        return None
    texto = texto.replace(",", "")
    try:
        return float(texto)
    except ValueError:
        return None

def importar_tarifas_desde_excel(ruta="tarifas.xlsx"):
    from openpyxl import load_workbook
    import os

    if not os.path.exists(ruta):
        return 0, 0, ["No se encontró el archivo tarifas.xlsx"]

    wb = load_workbook(ruta)
    hoja = wb.active

    insertadas = 0
    saltadas = 0
    errores = []

    for i, fila in enumerate(hoja.iter_rows(min_row=1), start=1):
        try:
            codigo = fila[0].value
            nombre = fila[1].value
            precio = fila[2].value
            extra = fila[3].value

            if not codigo or not nombre:
                saltadas += 1
                continue

            codigo = int(codigo)

            if precio is None:
                saltadas += 1
                continue

            precio = float(precio)

            if extra is None:
                extra = 0
            else:
                extra = float(extra)

            guardar_tarifa_db(codigo, str(nombre), precio, extra)
            insertadas += 1

        except Exception as e:
            errores.append(f"Fila {i}: {str(e)}")
            saltadas += 1

    return insertadas, saltadas, errores

def importar_tarifas_desde_txt(ruta_archivo="tarifas.txt"):
    import os

    if not os.path.exists(ruta_archivo):
        return 0, 0, [f"No existe el archivo: {ruta_archivo}"]

    insertadas = 0
    saltadas = 0
    errores = []

    with open(ruta_archivo, "r", encoding="utf-8") as f:
        lineas = f.readlines()

    for num_linea, linea in enumerate(lineas, start=1):
        linea = linea.strip()

        if not linea:
            continue

        partes = [p.strip() for p in linea.split("\t") if p.strip() != ""]

        if len(partes) < 4:
            saltadas += 1
            errores.append(f"Línea {num_linea}: formato incompleto -> {linea}")
            continue

        codigo_txt = partes[0]
        nombre = partes[1]
        precio_txt = partes[2]
        extra_txt = partes[3]

        try:
            codigo = int(codigo_txt)
        except ValueError:
            saltadas += 1
            errores.append(f"Línea {num_linea}: código inválido -> {codigo_txt}")
            continue

        precio = limpiar_numero(precio_txt)
        extra = limpiar_numero(extra_txt)

        if precio is None:
            saltadas += 1
            errores.append(f"Línea {num_linea}: precio inválido -> {precio_txt}")
            continue

        if extra is None:
            extra = 0

        if not nombre:
            saltadas += 1
            errores.append(f"Línea {num_linea}: nombre vacío")
            continue

        guardar_tarifa_db(codigo, nombre, precio, extra)
        insertadas += 1

    return insertadas, saltadas, errores


def obtener_tarifa_por_codigo(codigo):
    if usar_backend_nube():
        try:
            data = api_get_json(f"/tarifas/{int(codigo)}")
            if data and not data.get("error"):
                return (
                    data.get("codigo"),
                    data.get("nombre"),
                    data.get("precio", 0),
                    data.get("extra", 0),
                )
        except Exception:
            pass
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT codigo, nombre, precio, extra
        FROM tarifas
        WHERE codigo = ?
    """, (codigo,))
    fila = cur.fetchone()
    conn.close()
    return fila


def eliminar_tarifa_por_codigo(codigo):
    if usar_backend_nube():
        data = api_delete_json(f"/tarifas/{int(codigo)}")
        return 1 if data and data.get("ok") else 0
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM tarifas WHERE codigo = ?", (codigo,))
    cambios = cur.rowcount
    conn.commit()
    conn.close()
    return cambios


def obtener_todas_tarifas():
    if usar_backend_nube():
        try:
            data = api_get_json("/tarifas")
            if isinstance(data, list):
                return [
                    (
                        item.get("codigo"),
                        item.get("nombre"),
                        item.get("precio", 0),
                        item.get("extra", 0),
                    )
                    for item in data
                ]
        except Exception:
            pass
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT codigo, nombre, precio, extra
        FROM tarifas
        ORDER BY codigo
    """)
    filas = cur.fetchall()
    conn.close()
    return filas

def obtener_todos_los_documentos(filtro=""):
    conn = conectar_db()
    cur = conn.cursor()

    filtro = f"%{filtro.strip().lower()}%"

    cur.execute("""
        SELECT d.id, d.tipo, d.numero_doc, c.nombre, d.fecha, d.total_final, d.cerrado
        FROM documentos d
        JOIN clientes c ON d.cliente_id = c.id
        WHERE
            LOWER(c.nombre) LIKE ?
            OR LOWER(d.tipo) LIKE ?
            OR CAST(d.numero_doc AS TEXT) LIKE ?
            OR LOWER(d.fecha) LIKE ?
        ORDER BY d.id DESC
    """, (filtro, filtro, filtro, filtro))

    filas = cur.fetchall()
    conn.close()
    return filas

def borrar_todas_las_tarifas():
    if usar_backend_nube():
        api_delete_json("/tarifas")
        return
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM tarifas")
    conn.commit()
    conn.close()


# =========================
# DOCUMENTOS
# =========================
def siguiente_numero_documento(tipo):
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT COALESCE(MAX(numero_doc), 0)
        FROM documentos
        WHERE tipo = ?
    """, (tipo,))
    ultimo = cur.fetchone()[0] or 0
    conn.close()
    return ultimo + 1

def crear_documento_db(tipo, cliente_id, fecha, fecha_entrega=""):
    if usar_backend_nube():
        data = api_post_json("/documentos", {
            "tipo": tipo,
            "cliente_id": cliente_id,
            "fecha": fecha,
            "fecha_entrega": fecha_entrega or "",
        })
        return data.get("id") if isinstance(data, dict) else None
    numero_doc = siguiente_numero_documento(tipo)

    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO documentos (
            tipo, numero_doc, cliente_id, fecha, fecha_entrega,
            subtotal, descuento, itbis, total_final, cerrado
        )
        VALUES (?, ?, ?, ?, ?, 0, 0, 0, 0, 0)
    """, (tipo, numero_doc, cliente_id, fecha, fecha_entrega or ""))
    documento_id = cur.lastrowid
    conn.commit()
    conn.close()
    return documento_id


def obtener_documento(documento_id):
    if usar_backend_nube():
        try:
            data = api_get_json(f"/documentos/{documento_id}")
            if data and not data.get("error"):
                cliente = data.get("cliente") or {}
                return (
                    data.get("id"),
                    data.get("tipo"),
                    data.get("numero_doc"),
                    data.get("cliente_id"),
                    cliente.get("nombre"),
                    cliente.get("telefono"),
                    cliente.get("rnc"),
                    cliente.get("direccion"),
                    data.get("fecha"),
                    data.get("subtotal", 0),
                    data.get("descuento", 0),
                    data.get("itbis", 0),
                    data.get("total_final", 0),
                    1 if data.get("cerrado") else 0,
                    data.get("metodo_pago"),
                    1 if data.get("retirado") else 0,
                    data.get("fecha_entrega", ""),
                )
        except Exception:
            pass
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT d.id, d.tipo, d.numero_doc, d.cliente_id,
               c.nombre, c.telefono, c.rnc, c.direccion,
               d.fecha, d.subtotal, d.descuento, d.itbis, d.total_final, d.cerrado, d.metodo_pago, d.retirado, d.fecha_entrega
        FROM documentos d
        JOIN clientes c ON d.cliente_id = c.id
        WHERE d.id = ?
    """, (documento_id,))
    fila = cur.fetchone()
    conn.close()
    return fila


def obtener_documento_por_numero(numero_doc):
    if usar_backend_nube():
        try:
            data = api_get_json(f"/documentos/by-number/{numero_doc}")
            if data and not data.get("error"):
                cliente = data.get("cliente") or {}
                return (
                    data.get("id"),
                    data.get("tipo"),
                    data.get("numero_doc"),
                    data.get("cliente_id"),
                    cliente.get("nombre"),
                    cliente.get("telefono"),
                    cliente.get("rnc"),
                    cliente.get("direccion"),
                    data.get("fecha"),
                    data.get("subtotal", 0),
                    data.get("descuento", 0),
                    data.get("itbis", 0),
                    data.get("total_final", 0),
                    1 if data.get("cerrado") else 0,
                    data.get("metodo_pago"),
                    1 if data.get("retirado") else 0,
                    data.get("fecha_entrega", ""),
                )
        except Exception:
            pass
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT d.id, d.tipo, d.numero_doc, d.cliente_id,
               c.nombre, c.telefono, c.rnc, c.direccion,
               d.fecha, d.subtotal, d.descuento, d.itbis, d.total_final, d.cerrado, d.metodo_pago, d.retirado, d.fecha_entrega
        FROM documentos d
        JOIN clientes c ON d.cliente_id = c.id
        WHERE d.numero_doc = ?
        ORDER BY d.id DESC
        LIMIT 1
    """, (numero_doc,))
    fila = cur.fetchone()
    conn.close()
    return fila

def obtener_todos_los_documentos(filtro=""):
    if usar_backend_nube():
        try:
            data = api_get_json("/documentos", params={"filtro": filtro, "limit": 300})
            if isinstance(data, list):
                filas = []
                for item in data:
                    cliente = item.get("cliente") or {}
                    filas.append(
                        (
                            item.get("id"),
                            item.get("tipo"),
                            item.get("numero_doc"),
                            cliente.get("nombre"),
                            item.get("fecha"),
                            item.get("total_final", 0),
                            1 if item.get("cerrado") else 0,
                        )
                    )
                return filas
        except Exception:
            pass
    conn = conectar_db()
    cur = conn.cursor()

    filtro = f"%{filtro.strip().lower()}%"

    cur.execute("""
        SELECT d.id, d.tipo, d.numero_doc, c.nombre, d.fecha, d.total_final, d.cerrado
        FROM documentos d
        JOIN clientes c ON d.cliente_id = c.id
        WHERE
            LOWER(c.nombre) LIKE ?
            OR LOWER(d.tipo) LIKE ?
            OR CAST(d.numero_doc AS TEXT) LIKE ?
            OR LOWER(d.fecha) LIKE ?
        ORDER BY d.id DESC
    """, (filtro, filtro, filtro, filtro))

    filas = cur.fetchall()
    conn.close()
    return filas


def cerrar_documento_db(documento_id):
    if usar_backend_nube():
        api_put_json(f"/documentos/{documento_id}", {"cerrado": True})
        return
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE documentos
        SET cerrado = 1
        WHERE id = ?
    """, (documento_id,))
    conn.commit()
    conn.close()


def eliminar_documento_db(documento_id):
    if usar_backend_nube():
        data = api_delete_json(f"/documentos/{documento_id}")
        return 1 if isinstance(data, dict) and data.get("ok") else 0
    conn = conectar_db()
    cur = conn.cursor()

    tiene_documento_id = columna_existe(cur, "ordenes", "documento_id")
    tiene_factura_id = columna_existe(cur, "ordenes", "factura_id")

    if tiene_documento_id and tiene_factura_id:
        cur.execute("""
            SELECT id
            FROM ordenes
            WHERE COALESCE(documento_id, factura_id) = ?
        """, (documento_id,))
    elif tiene_documento_id:
        cur.execute("""
            SELECT id
            FROM ordenes
            WHERE documento_id = ?
        """, (documento_id,))
    else:
        cur.execute("""
            SELECT id
            FROM ordenes
            WHERE factura_id = ?
        """, (documento_id,))

    orden_ids = [fila[0] for fila in cur.fetchall()]

    for orden_id in orden_ids:
        cur.execute("DELETE FROM orden_detalles WHERE orden_id = ?", (orden_id,))

    if tiene_documento_id and tiene_factura_id:
        cur.execute("DELETE FROM ordenes WHERE COALESCE(documento_id, factura_id) = ?", (documento_id,))
    elif tiene_documento_id:
        cur.execute("DELETE FROM ordenes WHERE documento_id = ?", (documento_id,))
    else:
        cur.execute("DELETE FROM ordenes WHERE factura_id = ?", (documento_id,))

    cur.execute("DELETE FROM documentos WHERE id = ?", (documento_id,))
    cambios = cur.rowcount

    conn.commit()
    conn.close()
    return cambios


def actualizar_metodo_pago_documento(documento_id, metodo_pago):
    if usar_backend_nube():
        api_put_json(f"/documentos/{documento_id}", {"metodo_pago": normalizar_metodo_pago(metodo_pago)})
        return
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("UPDATE documentos SET metodo_pago = ? WHERE id = ?", (normalizar_metodo_pago(metodo_pago), documento_id))
    conn.commit()
    conn.close()


def actualizar_estado_retiro_documento(documento_id, retirado):
    if usar_backend_nube():
        api_put_json(f"/documentos/{documento_id}", {"retirado": bool(retirado)})
        return
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("UPDATE documentos SET retirado = ? WHERE id = ?", (1 if retirado else 0, documento_id))
    conn.commit()
    conn.close()


def actualizar_fecha_entrega_documento(documento_id, fecha_entrega):
    if usar_backend_nube():
        api_put_json(f"/documentos/{documento_id}", {"fecha_entrega": (fecha_entrega or "").strip()})
        return
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("UPDATE documentos SET fecha_entrega = ? WHERE id = ?", ((fecha_entrega or "").strip(), documento_id))
    conn.commit()
    conn.close()


def recalcular_estado_cobros_documento(documento_id):
    documento = obtener_documento(documento_id)
    if not documento:
        return
    total_final = float(documento[12] or 0)
    acumulado = obtener_total_cobrado_documento(documento_id)
    pagado_total = 1 if acumulado >= total_final and total_final > 0 else 0

    if usar_backend_nube():
        data = api_get_json(f"/documentos/{documento_id}")
        if not data or data.get("error"):
            return
        for cobro in data.get("cobros", []):
            api_put_json(f"/cobros/{cobro['id']}", {"pagado_total": bool(pagado_total)})
        return

    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("UPDATE cobros SET pagado_total = ? WHERE documento_id = ?", (pagado_total, documento_id))
    conn.commit()
    conn.close()


def obtener_cobro_automatico_facturacion(documento_id):
    if usar_backend_nube():
        data = api_get_json(f"/documentos/{documento_id}")
        if not data or data.get("error"):
            return None
        for cobro in reversed(data.get("cobros", [])):
            if (cobro.get("referencia") or "") == REFERENCIA_COBRO_AUTO:
                return (
                    cobro.get("id"),
                    cobro.get("fecha"),
                    cobro.get("monto"),
                    cobro.get("metodo_pago"),
                    cobro.get("referencia") or "",
                )
        return None
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, fecha, monto, metodo_pago, COALESCE(referencia, '')
        FROM cobros
        WHERE documento_id = ? AND referencia = ?
        ORDER BY id DESC
        LIMIT 1
    """, (documento_id, REFERENCIA_COBRO_AUTO))
    fila = cur.fetchone()
    conn.close()
    return fila


def sincronizar_cobro_automatico_facturacion(documento_id, fecha, monto, metodo_pago):
    if not documento_id:
        return

    monto = max(0.0, float(monto or 0))
    metodo_pago = normalizar_metodo_pago(metodo_pago)
    existente = obtener_cobro_automatico_facturacion(documento_id)

    if usar_backend_nube():
        if monto <= 0:
            if existente:
                api_delete_json(f"/cobros/{existente[0]}")
        elif existente:
            api_put_json(f"/cobros/{existente[0]}", {"fecha": fecha, "monto": monto, "metodo_pago": metodo_pago, "referencia": REFERENCIA_COBRO_AUTO})
        else:
            documento = obtener_documento(documento_id)
            if documento:
                api_post_json("/cobros", {
                    "documento_id": documento_id,
                    "numero_doc": documento[2],
                    "cliente_nombre": documento[4],
                    "fecha": fecha,
                    "monto": monto,
                    "metodo_pago": metodo_pago,
                    "referencia": REFERENCIA_COBRO_AUTO,
                    "pagado_total": False,
                })
        actualizar_metodo_pago_documento(documento_id, metodo_pago)
        recalcular_estado_cobros_documento(documento_id)
        return

    conn = conectar_db()
    cur = conn.cursor()
    if monto <= 0:
        cur.execute("DELETE FROM cobros WHERE documento_id = ? AND referencia = ?", (documento_id, REFERENCIA_COBRO_AUTO))
    elif existente:
        cur.execute("""
            UPDATE cobros
            SET fecha = ?, monto = ?, metodo_pago = ?
            WHERE id = ?
        """, (fecha, monto, metodo_pago, existente[0]))
    else:
        documento = obtener_documento(documento_id)
        if documento:
            cur.execute("""
                INSERT INTO cobros (documento_id, numero_doc, cliente_nombre, fecha, monto, metodo_pago, referencia, pagado_total)
                VALUES (?, ?, ?, ?, ?, ?, ?, 0)
            """, (documento_id, documento[2], documento[4], fecha, monto, metodo_pago, REFERENCIA_COBRO_AUTO))
    conn.commit()
    conn.close()

    actualizar_metodo_pago_documento(documento_id, metodo_pago)
    recalcular_estado_cobros_documento(documento_id)


def registrar_cobro_db(documento_id, fecha, monto, metodo_pago, referencia=""):
    documento = obtener_documento(documento_id)
    if not documento:
        return None

    numero_doc = documento[2]
    cliente_nombre = documento[4]
    total_final = float(documento[12] or 0)
    monto = max(0.0, float(monto or 0))
    metodo_pago = normalizar_metodo_pago(metodo_pago)
    acumulado = obtener_total_cobrado_documento(documento_id) + monto
    pagado_total = 1 if acumulado >= total_final and total_final > 0 else 0

    if usar_backend_nube():
        data = api_post_json("/cobros", {
            "documento_id": documento_id,
            "numero_doc": numero_doc,
            "cliente_nombre": cliente_nombre,
            "fecha": fecha,
            "monto": monto,
            "metodo_pago": metodo_pago,
            "referencia": referencia,
            "pagado_total": bool(pagado_total),
        })
        actualizar_metodo_pago_documento(documento_id, metodo_pago)
        recalcular_estado_cobros_documento(documento_id)
        return data.get("id") if isinstance(data, dict) else None

    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO cobros (documento_id, numero_doc, cliente_nombre, fecha, monto, metodo_pago, referencia, pagado_total)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (documento_id, numero_doc, cliente_nombre, fecha, monto, metodo_pago, referencia, pagado_total))
    cobro_id = cur.lastrowid
    conn.commit()
    conn.close()

    actualizar_metodo_pago_documento(documento_id, metodo_pago)
    return cobro_id


def obtener_cobro_por_id(cobro_id):
    if usar_backend_nube():
        docs = api_get_json("/documentos", params={"limit": 300}) or []
        for doc in docs:
            for cobro in doc.get("cobros", []):
                if cobro.get("id") == cobro_id:
                    return (
                        cobro.get("id"),
                        doc.get("id"),
                        cobro.get("numero_doc", doc.get("numero_doc")),
                        cobro.get("cliente_nombre", (doc.get("cliente") or {}).get("nombre")),
                        cobro.get("fecha"),
                        cobro.get("monto", 0),
                        cobro.get("metodo_pago"),
                        cobro.get("referencia") or "",
                        1 if cobro.get("pagado_total") else 0,
                    )
        return None
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, documento_id, numero_doc, cliente_nombre, fecha, monto, metodo_pago, COALESCE(referencia, ''), pagado_total
        FROM cobros
        WHERE id = ?
    """, (cobro_id,))
    fila = cur.fetchone()
    conn.close()
    return fila


def actualizar_cobro_db(cobro_id, fecha, monto, metodo_pago, referencia=""):
    cobro = obtener_cobro_por_id(cobro_id)
    if not cobro:
        return 0

    documento_id = cobro[1]
    documento = obtener_documento(documento_id)
    if not documento:
        return 0

    metodo_pago = normalizar_metodo_pago(metodo_pago)
    monto = max(0.0, float(monto or 0))

    if usar_backend_nube():
        data = api_put_json(f"/cobros/{cobro_id}", {
            "fecha": fecha,
            "monto": monto,
            "metodo_pago": metodo_pago,
            "referencia": referencia,
        })
        if isinstance(data, dict) and not data.get("error"):
            actualizar_metodo_pago_documento(documento_id, metodo_pago)
            recalcular_estado_cobros_documento(documento_id)
            return 1
        return 0

    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE cobros
        SET fecha = ?, monto = ?, metodo_pago = ?, referencia = ?
        WHERE id = ?
    """, (fecha, monto, metodo_pago, referencia, cobro_id))

    conn.commit()
    conn.close()
    actualizar_metodo_pago_documento(documento_id, metodo_pago)
    recalcular_estado_cobros_documento(documento_id)
    return 1


def obtener_cobros_documento(documento_id):
    if usar_backend_nube():
        data = api_get_json(f"/documentos/{documento_id}")
        if not data or data.get("error"):
            return []
        return [
            (
                item.get("id"),
                item.get("fecha"),
                item.get("monto", 0),
                item.get("metodo_pago"),
                item.get("referencia") or "",
                1 if item.get("pagado_total") else 0,
            )
            for item in data.get("cobros", [])
        ]
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, fecha, monto, metodo_pago, COALESCE(referencia, ''), pagado_total
        FROM cobros
        WHERE documento_id = ?
        ORDER BY id
    """, (documento_id,))
    filas = cur.fetchall()
    conn.close()
    return filas


def obtener_total_cobrado_documento(documento_id):
    if usar_backend_nube():
        return sum(float(c[2] or 0) for c in obtener_cobros_documento(documento_id))
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("SELECT COALESCE(SUM(monto), 0) FROM cobros WHERE documento_id = ?", (documento_id,))
    total = cur.fetchone()[0] or 0
    conn.close()
    return float(total)


def obtener_todos_los_cobros(filtro=""):
    conn = conectar_db()
    cur = conn.cursor()
    filtro = f"%{(filtro or '').strip().lower()}%"
    cur.execute("""
        SELECT id, numero_doc, cliente_nombre, fecha, monto, metodo_pago, COALESCE(referencia, ''), pagado_total
        FROM cobros
        WHERE
            LOWER(cliente_nombre) LIKE ?
            OR CAST(numero_doc AS TEXT) LIKE ?
            OR LOWER(metodo_pago) LIKE ?
            OR LOWER(fecha) LIKE ?
            OR LOWER(COALESCE(referencia, '')) LIKE ?
        ORDER BY id DESC
    """, (filtro, filtro, filtro, filtro, filtro))
    filas = cur.fetchall()
    conn.close()
    return filas


def obtener_resumen_documentos_cobro(filtro=""):
    if usar_backend_nube():
        try:
            docs = api_get_json("/documentos", params={"filtro": filtro, "limit": 300}) or []
            filas = []
            for item in docs:
                cliente = item.get("cliente") or {}
                cobrado = sum(float(c.get("monto", 0) or 0) for c in item.get("cobros", []))
                filas.append((
                    item.get("id"),
                    item.get("numero_doc"),
                    cliente.get("nombre"),
                    item.get("fecha"),
                    item.get("total_final", 0),
                    item.get("metodo_pago"),
                    1 if item.get("retirado") else 0,
                    cobrado,
                ))
            return filas
        except Exception:
            pass
    conn = conectar_db()
    cur = conn.cursor()
    filtro = f"%{(filtro or '').strip().lower()}%"
    cur.execute("""
        SELECT d.id, d.numero_doc, c.nombre, d.fecha, d.total_final, d.metodo_pago, d.retirado,
               COALESCE((SELECT SUM(monto) FROM cobros cb WHERE cb.documento_id = d.id), 0)
        FROM documentos d
        JOIN clientes c ON d.cliente_id = c.id
        WHERE
            LOWER(c.nombre) LIKE ?
            OR CAST(d.numero_doc AS TEXT) LIKE ?
            OR LOWER(d.fecha) LIKE ?
            OR LOWER(d.tipo) LIKE ?
        ORDER BY d.id DESC
    """, (filtro, filtro, filtro, filtro))
    filas = cur.fetchall()
    conn.close()
    return filas


def actualizar_totales_documento(documento_id, descuento=0):
    if usar_backend_nube():
        data = api_get_json(f"/documentos/{documento_id}")
        if not data or data.get("error"):
            return 0, 0, 0, 0
        subtotal = sum(float((orden or {}).get("total_orden", 0) or 0) for orden in data.get("ordenes", []))
        resumen = calcular_resumen_financiero(subtotal, descuento)
        api_put_json(f"/documentos/{documento_id}", {
            "subtotal": resumen["subtotal"],
            "descuento": resumen["descuento_pct"],
            "itbis": resumen["itbis"],
            "total_final": resumen["total_final"],
        })
        return (
            resumen["subtotal"],
            resumen["descuento_pct"],
            resumen["itbis"],
            resumen["total_final"],
        )

    conn = conectar_db()
    cur = conn.cursor()

    tiene_documento_id = columna_existe(cur, "ordenes", "documento_id")
    tiene_factura_id = columna_existe(cur, "ordenes", "factura_id")

    if tiene_documento_id and tiene_factura_id:
        cur.execute("""
            SELECT COALESCE(SUM(total_orden), 0)
            FROM ordenes
            WHERE COALESCE(documento_id, factura_id) = ?
        """, (documento_id,))
    elif tiene_documento_id:
        cur.execute("""
            SELECT COALESCE(SUM(total_orden), 0)
            FROM ordenes
            WHERE documento_id = ?
        """, (documento_id,))
    else:
        cur.execute("""
            SELECT COALESCE(SUM(total_orden), 0)
            FROM ordenes
            WHERE factura_id = ?
        """, (documento_id,))
    subtotal = cur.fetchone()[0] or 0
    resumen = calcular_resumen_financiero(subtotal, descuento)

    cur.execute("""
        UPDATE documentos
        SET subtotal = ?, descuento = ?, itbis = ?, total_final = ?
        WHERE id = ?
    """, (
        resumen["subtotal"],
        resumen["descuento_pct"],
        resumen["itbis"],
        resumen["total_final"],
        documento_id,
    ))

    conn.commit()
    conn.close()

    return (
        resumen["subtotal"],
        resumen["descuento_pct"],
        resumen["itbis"],
        resumen["total_final"],
    )


def crear_orden_db(documento_id, a_enmarcar, ancho, largo, total_orden, notas=""):
    if usar_backend_nube():
        data = api_post_json("/ordenes", {
            "documento_id": documento_id,
            "a_enmarcar": a_enmarcar,
            "notas": notas or "",
            "ancho": ancho,
            "largo": largo,
            "total_orden": total_orden,
        })
        return data.get("id") if isinstance(data, dict) else None
    conn = conectar_db()
    cur = conn.cursor()
    tiene_factura_id = columna_existe(cur, "ordenes", "factura_id")
    tiene_documento_id = columna_existe(cur, "ordenes", "documento_id")
    tiene_notas = columna_existe(cur, "ordenes", "notas")

    if tiene_factura_id and tiene_documento_id and tiene_notas:
        cur.execute("""
            INSERT INTO ordenes (factura_id, documento_id, a_enmarcar, notas, ancho, largo, total_orden)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (documento_id, documento_id, a_enmarcar, notas, ancho, largo, total_orden))
    elif tiene_documento_id and tiene_notas:
        cur.execute("""
            INSERT INTO ordenes (documento_id, a_enmarcar, notas, ancho, largo, total_orden)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (documento_id, a_enmarcar, notas, ancho, largo, total_orden))
    elif tiene_factura_id and tiene_documento_id:
        cur.execute("""
            INSERT INTO ordenes (factura_id, documento_id, a_enmarcar, ancho, largo, total_orden)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (documento_id, documento_id, a_enmarcar, ancho, largo, total_orden))
    elif tiene_documento_id:
        cur.execute("""
            INSERT INTO ordenes (documento_id, a_enmarcar, ancho, largo, total_orden)
            VALUES (?, ?, ?, ?, ?)
        """, (documento_id, a_enmarcar, ancho, largo, total_orden))
    else:
        cur.execute("""
            INSERT INTO ordenes (factura_id, a_enmarcar, ancho, largo, total_orden)
            VALUES (?, ?, ?, ?, ?)
        """, (documento_id, a_enmarcar, ancho, largo, total_orden))

    orden_id = cur.lastrowid
    conn.commit()
    conn.close()
    return orden_id


def guardar_detalle_orden_db(
    orden_id, cantidad, codigo_material, descripcion_material,
    ancho, largo, pies, precio, subtotal, total
):
    if usar_backend_nube():
        data = api_post_json("/orden-detalles", {
            "orden_id": orden_id,
            "cantidad": cantidad,
            "codigo_material": codigo_material,
            "descripcion_material": descripcion_material,
            "ancho": ancho,
            "largo": largo,
            "pies": pies,
            "precio": precio,
            "subtotal": subtotal,
            "total": total,
        })
        return data.get("id") if isinstance(data, dict) else None
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO orden_detalles (
            orden_id, cantidad, codigo_material, descripcion_material,
            ancho, largo, pies, precio, subtotal, total
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        orden_id, cantidad, codigo_material, descripcion_material,
        ancho, largo, pies, precio, subtotal, total
    ))
    conn.commit()
    conn.close()


def obtener_ordenes_de_documento(documento_id):
    conn = conectar_db()
    cur = conn.cursor()
    tiene_documento_id = columna_existe(cur, "ordenes", "documento_id")
    tiene_factura_id = columna_existe(cur, "ordenes", "factura_id")

    if tiene_documento_id and tiene_factura_id:
        cur.execute("""
            SELECT id, a_enmarcar, COALESCE(notas, ''), ancho, largo, total_orden
            FROM ordenes
            WHERE COALESCE(documento_id, factura_id) = ?
            ORDER BY id
        """, (documento_id,))
    elif tiene_documento_id:
        cur.execute("""
            SELECT id, a_enmarcar, COALESCE(notas, ''), ancho, largo, total_orden
            FROM ordenes
            WHERE documento_id = ?
            ORDER BY id
        """, (documento_id,))
    else:
        cur.execute("""
            SELECT id, a_enmarcar, '' as notas, ancho, largo, total_orden
            FROM ordenes
            WHERE factura_id = ?
            ORDER BY id
        """, (documento_id,))
    filas = cur.fetchall()
    conn.close()
    return filas


# =========================
# CÁLCULOS
# =========================
def calcular_pies(codigo, ancho, largo, extra):
    if 3000 <= codigo <= 80000:
        return (extra + 2 * (ancho + largo)) / 12
    return ancho * largo


# =========================
# APP
# =========================
class App(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Sistema de Facturación PF")
        self.geometry("1600x900")

        crear_tablas()

        self.documento_actual_id = None
        self.tipo_actual = None

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # menú lateral
        self.menu = ctk.CTkFrame(self, width=220, corner_radius=0)
        self.menu.grid(row=0, column=0, sticky="ns")
        self.menu.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            self.menu,
            text="ENMARCADOS PF",
            font=ctk.CTkFont(size=24, weight="bold")
        ).grid(row=0, column=0, padx=20, pady=(20, 30), sticky="w")

        ctk.CTkButton(self.menu, text="CLIENTES", command=self.ir_clientes).grid(
            row=1, column=0, padx=20, pady=8, sticky="ew"
        )

        ctk.CTkButton(self.menu, text="TARIFAS", command=self.ir_tarifas).grid(
            row=2, column=0, padx=20, pady=8, sticky="ew"
        )

        ctk.CTkButton(self.menu, text="ORDEN", command=self.ir_orden).grid(
            row=3, column=0, padx=20, pady=8, sticky="ew"
        )

        ctk.CTkButton(self.menu, text="FACTURACIÓN", command=self.ir_facturacion).grid(
            row=4, column=0, padx=20, pady=8, sticky="ew"
        )

        ctk.CTkButton(self.menu, text="HISTORIAL", command=self.ir_historial).grid(
            row=5, column=0, padx=20, pady=8, sticky="ew"
        )

        ctk.CTkButton(self.menu, text="GESTIÓN DE COBRO", command=self.ir_cobros).grid(
            row=6, column=0, padx=20, pady=8, sticky="ew"
        )

        ctk.CTkButton(self.menu, text="BACKUP", command=self.ir_backup).grid(
            row=7, column=0, padx=20, pady=8, sticky="ew"
        )

        # área de contenido
        self.content = ctk.CTkFrame(self, corner_radius=0)
        self.content.grid(row=0, column=1, sticky="nsew")
        self.content.grid_rowconfigure(0, weight=1)
        self.content.grid_columnconfigure(0, weight=1)

        # páginas
        self.pagina_clientes = PaginaClientes(self.content)
        self.pagina_tarifas = PaginaTarifas(self.content)
        self.pagina_orden = PaginaOrden(self)
        self.pagina_facturacion = PaginaFacturacion(self)
        self.pagina_historial = PaginaHistorial(self)
        self.pagina_cobros = PaginaCobros(self)
        self.pagina_backup = PaginaBackup(self)

        self.pagina_clientes.grid(row=0, column=0, sticky="nsew")
        self.pagina_tarifas.grid(row=0, column=0, sticky="nsew")
        self.pagina_orden.grid(row=0, column=0, sticky="nsew")
        self.pagina_facturacion.grid(row=0, column=0, sticky="nsew")
        self.pagina_historial.grid(row=0, column=0, sticky="nsew")
        self.pagina_cobros.grid(row=0, column=0, sticky="nsew")
        self.pagina_backup.grid(row=0, column=0, sticky="nsew")

        self.ir_facturacion()

    def ir_clientes(self):
        self.pagina_clientes.cargar_lista_clientes()
        self.pagina_clientes.tkraise()

    def ir_tarifas(self):
        dialogo = ctk.CTkInputDialog(
            text="Digite el código para entrar a TARIFAS",
            title="Acceso"
        )
        codigo = dialogo.get_input()

        if codigo != "2005":
            messagebox.showerror("Acceso denegado", "Código incorrecto.")
            return

        self.pagina_tarifas.cargar_lista_tarifas()
        self.pagina_tarifas.tkraise()

    def ir_orden(self):
        self.pagina_orden.refrescar_desde_facturacion()
        self.pagina_orden.tkraise()

    def ir_facturacion(self):
        self.pagina_facturacion.refrescar()
        self.pagina_facturacion.tkraise()

    def ir_historial(self):
        self.pagina_historial.cargar_documentos()
        self.pagina_historial.tkraise()

    def ir_cobros(self):
        self.pagina_cobros.cargar_cobros()
        self.pagina_cobros.tkraise()

    def ir_backup(self):
        self.pagina_backup.tkraise()

# =========================
# CLIENTES
# =========================
class PaginaClientes(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)

        self.configure(fg_color="#F3F7F7")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(3, weight=1)

        header = ctk.CTkFrame(self, fg_color=COLOR_PANEL, border_width=1, border_color=COLOR_BORDE, corner_radius=18)
        header.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="ew")
        header.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            header,
            text="CLIENTES",
            font=ctk.CTkFont(size=26, weight="bold")
        ).grid(row=0, column=0, padx=20, pady=(16, 2), sticky="w")

        ctk.CTkLabel(
            header,
            text="Busca, edita y guarda desde el mismo formulario",
            font=ctk.CTkFont(size=13),
            text_color=COLOR_TEXTO_SUAVE
        ).grid(row=1, column=0, padx=20, pady=(0, 16), sticky="w")

        # formulario
        form = ctk.CTkFrame(self, fg_color="white", border_width=1, border_color=COLOR_BORDE, corner_radius=18)
        form.grid(row=1, column=0, padx=20, pady=10, sticky="ew")

        for i in range(5):
            form.grid_columnconfigure(i, weight=1)

        self.nombre = ctk.CTkEntry(form, placeholder_text="Nombre")
        self.nombre.grid(row=0, column=0, padx=8, pady=8, sticky="ew")
        self.nombre.bind("<FocusOut>", lambda _e: self.cargar_por_nombre(silencioso=True))

        self.telefono = ctk.CTkEntry(form, placeholder_text="Teléfono")
        self.telefono.grid(row=0, column=1, padx=8, pady=8, sticky="ew")

        self.rnc = ctk.CTkEntry(form, placeholder_text="RNC")
        self.rnc.grid(row=0, column=2, padx=8, pady=8, sticky="ew")

        self.direccion = ctk.CTkEntry(form, placeholder_text="Dirección")
        self.direccion.grid(row=0, column=3, padx=8, pady=8, sticky="ew")

        acciones = ctk.CTkFrame(form, fg_color="transparent")
        acciones.grid(row=0, column=4, rowspan=2, padx=8, pady=8, sticky="nsew")
        acciones.grid_columnconfigure((0, 1), weight=1)

        ctk.CTkButton(form, text="Guardar", command=self.guardar, fg_color=COLOR_PRINCIPAL, hover_color=COLOR_PRINCIPAL_HOVER, text_color="black").grid(
            row=1, column=0, padx=8, pady=(0, 10), sticky="ew"
        )

        ctk.CTkButton(form, text="Nuevo", command=self.limpiar_formulario, fg_color="white", hover_color="#f2f2f2", text_color="black", border_width=1, border_color=COLOR_BORDE).grid(
            row=1, column=1, padx=8, pady=(0, 10), sticky="ew"
        )

        ctk.CTkButton(acciones, text="Cargar", command=self.cargar_por_nombre, fg_color="white", hover_color="#f2f2f2", text_color="black", border_width=1, border_color=COLOR_BORDE).grid(
            row=0, column=0, padx=4, pady=4, sticky="ew"
        )

        ctk.CTkButton(acciones, text="Importar", command=self.importar_excel, fg_color="white", hover_color="#f2f2f2", text_color="black", border_width=1, border_color=COLOR_BORDE).grid(
            row=0, column=1, padx=4, pady=4, sticky="ew"
        )

        ctk.CTkButton(acciones, text="Eliminar", command=self.eliminar_por_nombre, fg_color="black", hover_color="#222222", text_color="white").grid(
            row=1, column=0, columnspan=2, padx=4, pady=4, sticky="ew"
        )

        # buscador
        buscador = ctk.CTkFrame(self, fg_color="white", border_width=1, border_color=COLOR_BORDE, corner_radius=18)
        buscador.grid(row=2, column=0, padx=20, pady=(0, 10), sticky="ew")
        buscador.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(buscador, text="Buscar:").grid(row=0, column=0, padx=8, pady=8)
        self.buscar_entry = ctk.CTkEntry(
            buscador,
            placeholder_text="Nombre, teléfono, RNC o dirección"
        )
        self.buscar_entry.grid(row=0, column=1, padx=8, pady=8, sticky="ew")
        self.buscar_entry.bind("<KeyRelease>", self.filtrar_lista)

        ctk.CTkButton(buscador, text="Eliminar seleccionados", command=self.eliminar_seleccionados, fg_color="white", hover_color="#f2f2f2", text_color="black", border_width=1, border_color=COLOR_BORDE).grid(
            row=0, column=2, padx=8, pady=8
        )

        # lista ligera
        lista_wrap = ctk.CTkFrame(self, fg_color="white", border_width=1, border_color=COLOR_BORDE, corner_radius=18)
        lista_wrap.grid(row=3, column=0, padx=20, pady=15, sticky="nsew")
        lista_wrap.grid_rowconfigure(0, weight=1)
        lista_wrap.grid_columnconfigure(0, weight=1)

        columnas = ("nombre", "telefono", "rnc", "direccion")
        self.lista_clientes = ttk.Treeview(lista_wrap, columns=columnas, show="headings", selectmode="extended")
        self.lista_clientes.heading("nombre", text="Nombre")
        self.lista_clientes.heading("telefono", text="Teléfono")
        self.lista_clientes.heading("rnc", text="RNC")
        self.lista_clientes.heading("direccion", text="Dirección")
        self.lista_clientes.column("nombre", width=260, anchor="w")
        self.lista_clientes.column("telefono", width=140, anchor="w")
        self.lista_clientes.column("rnc", width=140, anchor="w")
        self.lista_clientes.column("direccion", width=360, anchor="w")
        self.lista_clientes.grid(row=0, column=0, sticky="nsew", padx=(12, 0), pady=12)
        self.lista_clientes.bind("<<TreeviewSelect>>", self.seleccion_desde_lista_clientes)

        scroll_y = ttk.Scrollbar(lista_wrap, orient="vertical", command=self.lista_clientes.yview)
        scroll_y.grid(row=0, column=1, sticky="ns", padx=(0, 12), pady=12)
        self.lista_clientes.configure(yscrollcommand=scroll_y.set)

        self.cargar_lista_clientes()

    def limpiar_formulario(self):
        self.nombre.delete(0, "end")
        self.telefono.delete(0, "end")
        self.rnc.delete(0, "end")
        self.direccion.delete(0, "end")

    def guardar(self):
        nombre = self.nombre.get().strip()
        telefono = self.telefono.get().strip()
        rnc = self.rnc.get().strip()
        direccion = self.direccion.get().strip()

        if not nombre:
            messagebox.showerror("Error", "Debes escribir el nombre del cliente.")
            return

        guardar_cliente_db(nombre, telefono, rnc, direccion)
        self.cargar_lista_clientes()
        messagebox.showinfo("Éxito", "Cliente guardado o actualizado.")
        self.limpiar_formulario()

    def cargar_por_nombre(self, silencioso=False):
        nombre = self.nombre.get().strip()

        if not nombre:
            if not silencioso:
                messagebox.showerror("Error", "Escribe el nombre del cliente.")
            return

        cliente = obtener_cliente_por_nombre(nombre)
        if not cliente:
            if not silencioso:
                messagebox.showwarning("No encontrado", "No existe un cliente con ese nombre.")
            return

        _, nombre, telefono, rnc, direccion = cliente

        self.nombre.delete(0, "end")
        self.nombre.insert(0, nombre)

        self.telefono.delete(0, "end")
        self.telefono.insert(0, telefono or "")

        self.rnc.delete(0, "end")
        self.rnc.insert(0, rnc or "")

        self.direccion.delete(0, "end")
        self.direccion.insert(0, direccion or "")

    def eliminar_por_nombre(self):
        nombre = self.nombre.get().strip()

        if not nombre:
            messagebox.showerror("Error", "Escribe el nombre del cliente.")
            return

        respuesta = messagebox.askyesno(
            "Confirmar",
            f"¿Seguro que quieres eliminar el cliente '{nombre}'?"
        )
        if not respuesta:
            return

        cambios = eliminar_cliente_por_nombre(nombre)
        if cambios == 0:
            messagebox.showwarning("No encontrado", "No existe un cliente con ese nombre.")
            return

        self.cargar_lista_clientes()
        self.limpiar_formulario()
        messagebox.showinfo("Eliminado", "Cliente eliminado correctamente.")

    def borrar_todos(self):
        respuesta = messagebox.askyesno(
            "Confirmar",
            "¿Seguro que quieres borrar TODOS los clientes?"
        )
        if not respuesta:
            return

        borrar_todos_los_clientes()
        self.cargar_lista_clientes()
        self.limpiar_formulario()
        messagebox.showinfo("Listo", "Todos los clientes fueron eliminados.")

    def importar_excel(self):
        insertados, saltados, errores = importar_clientes_desde_excel("clientes.xlsx")
        self.cargar_lista_clientes()

        mensaje = f"Importados: {insertados}\nSaltados: {saltados}"
        if errores:
            mensaje += "\n\nErrores:\n" + "\n".join(errores[:10])

        messagebox.showinfo("Importación Excel", mensaje)

    def cargar_lista_clientes(self, filtro=""):
        for item in self.lista_clientes.get_children():
            self.lista_clientes.delete(item)
        filtro = filtro.lower().strip()

        for _, nombre, telefono, rnc, direccion in obtener_todos_clientes():
            texto_busqueda = f"{nombre} {telefono} {rnc} {direccion}".lower()
            if filtro and filtro not in texto_busqueda:
                continue
            self.lista_clientes.insert("", "end", iid=str(nombre), values=(nombre, telefono, rnc, direccion))

    def cargar_cliente_en_formulario(self, nombre):
        self.nombre.delete(0, "end")
        self.nombre.insert(0, nombre)
        self.cargar_por_nombre(silencioso=False)

    def seleccion_desde_lista_clientes(self, _event=None):
        seleccion = self.lista_clientes.selection()
        if len(seleccion) != 1:
            return
        nombre = seleccion[0]
        self.cargar_cliente_en_formulario(nombre)

    def filtrar_lista(self, event=None):
        filtro = self.buscar_entry.get()
        self.cargar_lista_clientes(filtro)

    def eliminar_seleccionados(self):
        seleccionados = list(self.lista_clientes.selection())

        if not seleccionados:
            messagebox.showwarning("Aviso", "No seleccionaste ningún cliente.")
            return

        respuesta = messagebox.askyesno(
            "Confirmar",
            f"¿Eliminar {len(seleccionados)} clientes seleccionados?"
        )
        if not respuesta:
            return

        conn = conectar_db()
        cur = conn.cursor()

        for nombre in seleccionados:
            cur.execute("DELETE FROM clientes WHERE nombre = ?", (nombre,))

        conn.commit()
        conn.close()

        self.cargar_lista_clientes()
        messagebox.showinfo("Listo", "Clientes eliminados correctamente.")

# =========================
# TARIFAS
# =========================
class PaginaTarifas(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)

        self.configure(fg_color="#F3F7F7")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(3, weight=1)

        header = ctk.CTkFrame(self, fg_color=COLOR_PANEL, border_width=1, border_color=COLOR_BORDE, corner_radius=18)
        header.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="ew")
        header.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            header,
            text="TARIFAS",
            font=ctk.CTkFont(size=26, weight="bold")
        ).grid(row=0, column=0, padx=20, pady=(16, 2), sticky="w")

        ctk.CTkLabel(
            header,
            text="Edita una tarifa escribiendo el mismo código y guardando de nuevo",
            font=ctk.CTkFont(size=13),
            text_color=COLOR_TEXTO_SUAVE
        ).grid(row=1, column=0, padx=20, pady=(0, 16), sticky="w")

        # formulario
        form = ctk.CTkFrame(self, fg_color="white", border_width=1, border_color=COLOR_BORDE, corner_radius=18)
        form.grid(row=1, column=0, padx=20, pady=10, sticky="ew")

        for i in range(5):
            form.grid_columnconfigure(i, weight=1)

        self.codigo = ctk.CTkEntry(form, placeholder_text="Código")
        self.codigo.grid(row=0, column=0, padx=8, pady=8, sticky="ew")
        self.codigo.bind("<FocusOut>", lambda _e: self.cargar_por_codigo(silencioso=True))

        self.nombre = ctk.CTkEntry(form, placeholder_text="Nombre")
        self.nombre.grid(row=0, column=1, padx=8, pady=8, sticky="ew")

        self.precio = ctk.CTkEntry(form, placeholder_text="Precio")
        self.precio.grid(row=0, column=2, padx=8, pady=8, sticky="ew")

        self.extra = ctk.CTkEntry(form, placeholder_text="Extra")
        self.extra.grid(row=0, column=3, padx=8, pady=8, sticky="ew")

        acciones = ctk.CTkFrame(form, fg_color="transparent")
        acciones.grid(row=0, column=4, rowspan=2, padx=8, pady=8, sticky="nsew")
        acciones.grid_columnconfigure((0, 1), weight=1)

        ctk.CTkButton(form, text="Guardar", command=self.guardar, fg_color=COLOR_PRINCIPAL, hover_color=COLOR_PRINCIPAL_HOVER, text_color="black").grid(
            row=1, column=0, padx=8, pady=(0, 10), sticky="ew"
        )

        ctk.CTkButton(form, text="Nuevo", command=self.limpiar_formulario, fg_color="white", hover_color="#f2f2f2", text_color="black", border_width=1, border_color=COLOR_BORDE).grid(
            row=1, column=1, padx=8, pady=(0, 10), sticky="ew"
        )

        ctk.CTkButton(acciones, text="Cargar", command=self.cargar_por_codigo, fg_color="white", hover_color="#f2f2f2", text_color="black", border_width=1, border_color=COLOR_BORDE).grid(
            row=0, column=0, padx=4, pady=4, sticky="ew"
        )

        ctk.CTkButton(acciones, text="Importar", command=self.importar_excel, fg_color="white", hover_color="#f2f2f2", text_color="black", border_width=1, border_color=COLOR_BORDE).grid(
            row=0, column=1, padx=4, pady=4, sticky="ew"
        )

        ctk.CTkButton(acciones, text="Eliminar", command=self.eliminar_por_codigo, fg_color="black", hover_color="#222222", text_color="white").grid(
            row=1, column=0, columnspan=2, padx=4, pady=4, sticky="ew"
        )

        # buscador
        buscador = ctk.CTkFrame(self, fg_color="white", border_width=1, border_color=COLOR_BORDE, corner_radius=18)
        buscador.grid(row=2, column=0, padx=20, pady=(0, 10), sticky="ew")
        buscador.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(buscador, text="Buscar:").grid(row=0, column=0, padx=8, pady=8)
        self.buscar_entry = ctk.CTkEntry(buscador, placeholder_text="Código o nombre")
        self.buscar_entry.grid(row=0, column=1, padx=8, pady=8, sticky="ew")
        self.buscar_entry.bind("<KeyRelease>", self.filtrar_lista)

        ctk.CTkButton(buscador, text="Eliminar seleccionadas", command=self.eliminar_seleccionadas, fg_color="white", hover_color="#f2f2f2", text_color="black", border_width=1, border_color=COLOR_BORDE).grid(
            row=0, column=2, padx=8, pady=8
        )

        # lista ligera
        lista_wrap = ctk.CTkFrame(self, fg_color="white", border_width=1, border_color=COLOR_BORDE, corner_radius=18)
        lista_wrap.grid(row=3, column=0, padx=20, pady=15, sticky="nsew")
        lista_wrap.grid_rowconfigure(0, weight=1)
        lista_wrap.grid_columnconfigure(0, weight=1)

        columnas = ("codigo", "nombre", "precio", "extra")
        self.lista_tarifas = ttk.Treeview(lista_wrap, columns=columnas, show="headings", selectmode="extended")
        self.lista_tarifas.heading("codigo", text="Código")
        self.lista_tarifas.heading("nombre", text="Nombre")
        self.lista_tarifas.heading("precio", text="Precio")
        self.lista_tarifas.heading("extra", text="Extra")
        self.lista_tarifas.column("codigo", width=90, anchor="center")
        self.lista_tarifas.column("nombre", width=360, anchor="w")
        self.lista_tarifas.column("precio", width=120, anchor="e")
        self.lista_tarifas.column("extra", width=120, anchor="e")
        self.lista_tarifas.grid(row=0, column=0, sticky="nsew", padx=(12, 0), pady=12)
        self.lista_tarifas.bind("<<TreeviewSelect>>", self.seleccion_desde_lista_tarifas)

        scroll_y = ttk.Scrollbar(lista_wrap, orient="vertical", command=self.lista_tarifas.yview)
        scroll_y.grid(row=0, column=1, sticky="ns", padx=(0, 12), pady=12)
        self.lista_tarifas.configure(yscrollcommand=scroll_y.set)

        self.cargar_lista_tarifas()

    def limpiar_formulario(self):
        self.codigo.delete(0, "end")
        self.nombre.delete(0, "end")
        self.precio.delete(0, "end")
        self.extra.delete(0, "end")

    def guardar(self):
        try:
            codigo = int(self.codigo.get().strip())
            nombre = self.nombre.get().strip()
            precio = float(self.precio.get().strip())
            extra = float(self.extra.get().strip() or 0)

            if not nombre:
                messagebox.showerror("Error", "Debes escribir el nombre.")
                return

            guardar_tarifa_db(codigo, nombre, precio, extra)
            self.cargar_lista_tarifas()
            messagebox.showinfo("Éxito", "Tarifa guardada o actualizada.")
            self.limpiar_formulario()
        except ValueError:
            messagebox.showerror("Error", "Código, precio y extra deben ser numéricos.")

    def cargar_por_codigo(self, silencioso=False):
        try:
            codigo = int(self.codigo.get().strip())
        except ValueError:
            if not silencioso:
                messagebox.showerror("Error", "Escribe un código válido.")
            return

        tarifa = obtener_tarifa_por_codigo(codigo)
        if not tarifa:
            if not silencioso:
                messagebox.showwarning("No encontrada", "No existe una tarifa con ese código.")
            return

        _, nombre, precio, extra = tarifa

        self.nombre.delete(0, "end")
        self.nombre.insert(0, nombre)

        self.precio.delete(0, "end")
        self.precio.insert(0, str(precio))

        self.extra.delete(0, "end")
        self.extra.insert(0, str(extra))

    def eliminar_por_codigo(self):
        try:
            codigo = int(self.codigo.get().strip())
        except ValueError:
            messagebox.showerror("Error", "Escribe un código válido.")
            return

        respuesta = messagebox.askyesno(
            "Confirmar",
            f"¿Seguro que quieres eliminar la tarifa con código {codigo}?"
        )
        if not respuesta:
            return

        cambios = eliminar_tarifa_por_codigo(codigo)
        if cambios == 0:
            messagebox.showwarning("No encontrada", "No existe una tarifa con ese código.")
            return

        self.cargar_lista_tarifas()
        self.limpiar_formulario()
        messagebox.showinfo("Eliminada", "Tarifa eliminada correctamente.")

    def borrar_todas(self):
        respuesta = messagebox.askyesno(
            "Confirmar",
            "¿Seguro que quieres borrar TODAS las tarifas?"
        )
        if not respuesta:
            return

        borrar_todas_las_tarifas()
        self.cargar_lista_tarifas()
        self.limpiar_formulario()
        messagebox.showinfo("Listo", "Todas las tarifas fueron eliminadas.")

    def importar_excel(self):
        insertadas, saltadas, errores = importar_tarifas_desde_excel("tarifas.xlsx")
        self.cargar_lista_tarifas()

        mensaje = f"Importadas: {insertadas}\nSaltadas: {saltadas}"
        if errores:
            mensaje += "\n\nErrores:\n" + "\n".join(errores[:10])

        messagebox.showinfo("Importación Excel", mensaje)

        
    def cargar_lista_tarifas(self, filtro=""):
        for item in self.lista_tarifas.get_children():
            self.lista_tarifas.delete(item)
        filtro = filtro.lower().strip()

        for codigo, nombre, precio, extra in obtener_todas_tarifas():
            texto_busqueda = f"{codigo} {nombre}".lower()
            if filtro and filtro not in texto_busqueda:
                continue
            self.lista_tarifas.insert("", "end", iid=str(codigo), values=(codigo, nombre, f"{precio:.2f}", f"{extra:.2f}"))

    def cargar_tarifa_en_formulario(self, codigo):
        self.codigo.delete(0, "end")
        self.codigo.insert(0, str(codigo))
        self.cargar_por_codigo(silencioso=False)

    def seleccion_desde_lista_tarifas(self, _event=None):
        seleccion = self.lista_tarifas.selection()
        if len(seleccion) != 1:
            return
        self.cargar_tarifa_en_formulario(seleccion[0])

    def filtrar_lista(self, event=None):
        filtro = self.buscar_entry.get()
        self.cargar_lista_tarifas(filtro)

    def eliminar_seleccionadas(self):
        seleccionadas = [int(codigo) for codigo in self.lista_tarifas.selection()]

        if not seleccionadas:
            messagebox.showwarning("Aviso", "No seleccionaste ninguna tarifa.")
            return

        respuesta = messagebox.askyesno(
            "Confirmar",
            f"¿Eliminar {len(seleccionadas)} tarifas seleccionadas?"
        )
        if not respuesta:
            return

        conn = conectar_db()
        cur = conn.cursor()

        for codigo in seleccionadas:
            cur.execute("DELETE FROM tarifas WHERE codigo = ?", (codigo,))

        conn.commit()
        conn.close()

        self.cargar_lista_tarifas()
        messagebox.showinfo("Listo", "Tarifas eliminadas correctamente.")
# =========================
# ORDEN
# =========================
class PaginaOrden(ctk.CTkFrame):
    def __init__(self, app):
        super().__init__(app.content)
        self.app = app
        self.filas = []

        self.configure(fg_color="white")

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # ================= ENCABEZADO =================
        header = ctk.CTkFrame(self, fg_color="white", corner_radius=0)
        header.grid(row=0, column=0, sticky="ew", padx=20, pady=(20, 10))
        header.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            header,
            text="ENMARCADOS PF",
            font=ctk.CTkFont(size=30, weight="bold"),
            text_color="black"
        ).grid(row=0, column=0, sticky="w", padx=10)

        ctk.CTkLabel(
            header,
            text="ORDEN DE TRABAJO",
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color="black"
        ).grid(row=0, column=1, sticky="e", padx=10)

        # ================= DATOS CLIENTE =================
        top = ctk.CTkFrame(self, fg_color="white", border_width=2, border_color="#57C7C3")
        top.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 12))
        for i in range(6):
            top.grid_columnconfigure(i, weight=1)

        self.cliente_entry = ctk.CTkComboBox(
            top,
            values=self.get_nombres_clientes(),
            command=self.autocompletar_cliente_combo
        )
        self.cliente_entry.grid(row=0, column=0, columnspan=2, padx=8, pady=8, sticky="ew")
        self.cliente_entry.bind("<KeyRelease>", self.buscar_cliente_orden)
        self.cliente_entry.bind("<Return>", self.autocompletar_cliente_evento)
        self.cliente_entry.bind("<FocusOut>", self.autocompletar_cliente_evento)

        self.telefono_entry = ctk.CTkEntry(top, placeholder_text="Teléfono")
        self.telefono_entry.grid(row=0, column=2, padx=8, pady=8, sticky="ew")

        self.rnc_entry = ctk.CTkEntry(top, placeholder_text="RNC")
        self.rnc_entry.grid(row=0, column=3, padx=8, pady=8, sticky="ew")

        self.fecha_entry = ctk.CTkEntry(top, placeholder_text="Fecha")
        self.fecha_entry.grid(row=0, column=4, columnspan=2, padx=8, pady=8, sticky="ew")
        self.fecha_entry.insert(0, fecha_hoy_str())

        self.a_enmarcar_entry = ctk.CTkEntry(top, placeholder_text="Concepto general: qué se va a enmarcar")
        self.a_enmarcar_entry.grid(row=1, column=0, columnspan=4, padx=8, pady=(0, 8), sticky="ew")

        self.notas_entry = ctk.CTkEntry(top, placeholder_text="Notas rápidas del trabajo")
        self.notas_entry.grid(row=1, column=4, columnspan=2, padx=8, pady=(0, 8), sticky="ew")


        # ================= TABLA =================
        tabla_wrap = ctk.CTkFrame(self, fg_color="white", border_width=2, border_color="#57C7C3")
        tabla_wrap.grid(row=2, column=0, sticky="nsew", padx=20, pady=(0, 12))
        tabla_wrap.grid_columnconfigure(0, weight=1)
        tabla_wrap.grid_rowconfigure(0, weight=1)

        tabla = ctk.CTkScrollableFrame(tabla_wrap, fg_color="white")
        tabla.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)

        headers = [
            "CANTIDAD", "#", "DESCRIPCIÓN",
            "ANCHO", "LARGO", "PIES/P2", "PRECIO", "SUBTOTAL", "TOTAL"
        ]
        widths = [90, 90, 260, 80, 80, 95, 95, 110, 110]
        self.columnas_editables_orden = [0, 1, 3, 4]

        for col, text in enumerate(headers):
            ctk.CTkLabel(
                tabla,
                text=text,
                width=widths[col],
                height=32,
                fg_color="#57C7C3",
                text_color="white",
                corner_radius=6,
                font=ctk.CTkFont(size=13, weight="bold")
            ).grid(row=0, column=col, padx=2, pady=2, sticky="nsew")

        for row in range(1, 11):
            fila = []
            for col in range(len(headers)):
                entry = ctk.CTkEntry(
                    tabla,
                    width=widths[col],
                    fg_color="white",
                    text_color="black",
                    border_color="black"
                )
                entry.grid(row=row, column=col, padx=2, pady=2)

                if col in [2, 5, 6, 7, 8]:
                    entry.configure(state="readonly")

                fila.append(entry)

                if col in self.columnas_editables_orden:
                    entry.bind("<Return>", lambda e, r=row - 1, c=col: self.mover_foco_orden(r, c, "next"))
                    entry.bind("<Tab>", lambda e, r=row - 1, c=col: self.mover_foco_orden(r, c, "next"))
                    entry.bind("<Shift-Tab>", lambda e, r=row - 1, c=col: self.mover_foco_orden(r, c, "prev"))
                    entry.bind("<Right>", lambda e, r=row - 1, c=col: self.mover_foco_orden(r, c, "right"))
                    entry.bind("<Left>", lambda e, r=row - 1, c=col: self.mover_foco_orden(r, c, "left"))
                    entry.bind("<Down>", lambda e, r=row - 1, c=col: self.mover_foco_orden(r, c, "down"))
                    entry.bind("<Up>", lambda e, r=row - 1, c=col: self.mover_foco_orden(r, c, "up"))

            fila[0].bind("<KeyRelease>", lambda e, r=row - 1: self.calcular_fila(r))
            fila[1].bind("<KeyRelease>", lambda e, r=row - 1: self.calcular_fila(r))
            fila[3].bind("<KeyRelease>", lambda e, r=row - 1: self.calcular_fila(r))
            fila[4].bind("<KeyRelease>", lambda e, r=row - 1: self.calcular_fila(r))

            self.filas.append(fila)

        # ================= RESUMEN =================
        bottom = ctk.CTkFrame(self, fg_color="white", border_width=2, border_color="#57C7C3")
        bottom.grid(row=3, column=0, sticky="ew", padx=20, pady=(0, 20))
        for i in range(8):
            bottom.grid_columnconfigure(i, weight=1)

        ctk.CTkLabel(
            bottom,
            text="Notas del trabajo",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="black"
        ).grid(row=0, column=0, padx=8, pady=(12, 4), sticky="w")

        self.notas_texto = ctk.CTkTextbox(bottom, height=90, fg_color="white", text_color="black", border_color="black", border_width=1)
        self.notas_texto.grid(row=1, column=0, columnspan=4, padx=8, pady=(0, 12), sticky="ew")

        ctk.CTkLabel(
            bottom,
            text="TOTAL ORDEN",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color="black"
        ).grid(row=0, column=5, padx=8, pady=12, sticky="e")

        self.total_orden_entry = ctk.CTkEntry(
            bottom,
            width=160,
            fg_color="white",
            text_color="black",
            border_color="black"
        )
        self.total_orden_entry.grid(row=0, column=6, padx=8, pady=12, sticky="w")

        ctk.CTkButton(
            bottom,
            text="Calcular",
            fg_color="#57C7C3",
            hover_color="#3eb5b1",
            text_color="black",
            command=self.recalcular_todo
        ).grid(row=1, column=0, padx=8, pady=10, sticky="ew")

        ctk.CTkButton(
            bottom,
            text="Pasar orden a facturación",
            fg_color="black",
            hover_color="#222222",
            text_color="white",
            command=self.pasar_orden_a_facturacion
        ).grid(row=1, column=1, columnspan=3, padx=8, pady=10, sticky="ew")

        ctk.CTkButton(
            bottom,
            text="Limpiar",
            fg_color="white",
            hover_color="#f1f1f1",
            text_color="black",
            border_width=1,
            border_color="black",
            command=self.limpiar_orden
        ).grid(row=1, column=4, padx=8, pady=10, sticky="ew")

    # ================= CLIENTES =================
    def get_nombres_clientes(self):
        return filtrar_nombres_clientes("", limite=200)

    def refrescar_clientes(self):
        self.cliente_entry.configure(values=self.get_nombres_clientes())

    def buscar_cliente_orden(self, event=None):
        texto = self.cliente_entry.get().strip()
        sugerencias = filtrar_nombres_clientes(texto)
        self.cliente_entry.configure(values=sugerencias)

    def refrescar_desde_facturacion(self):
        self.refrescar_clientes()
        nombre = self.app.pagina_facturacion.cliente_combo.get().strip()
        if nombre:
            self.cliente_entry.set(nombre)
            self.autocompletar_cliente(nombre)
            fecha = self.app.pagina_facturacion.fecha_entry.get().strip()
            if fecha:
                self.fecha_entry.delete(0, "end")
                self.fecha_entry.insert(0, fecha)

    def autocompletar_cliente_evento(self, event=None):
        nombre = self.cliente_entry.get().strip()
        cliente = obtener_cliente_por_nombre(nombre)
        if not cliente:
            return

        _, _, telefono, rnc, _ = cliente

        self.telefono_entry.delete(0, "end")
        self.telefono_entry.insert(0, telefono or "")

        self.rnc_entry.delete(0, "end")
        self.rnc_entry.insert(0, rnc or "")

    def autocompletar_cliente_combo(self, nombre):
        self.autocompletar_cliente(nombre)

    def autocompletar_cliente_evento(self, event=None):
        nombre = self.cliente_entry.get().strip()
        self.autocompletar_cliente(nombre)

    def autocompletar_cliente(self, nombre):
        if not nombre:
            return

        cliente = obtener_cliente_por_nombre(nombre)
        if not cliente:
            return

        _, _, telefono, rnc, _ = cliente

        self.telefono_entry.delete(0, "end")
        self.telefono_entry.insert(0, telefono or "")

        self.rnc_entry.delete(0, "end")
        self.rnc_entry.insert(0, rnc or "")    
        self.sincronizar_cliente_a_facturacion(nombre)

    def sincronizar_cliente_a_facturacion(self, nombre):
        if not nombre:
            return
        self.app.pagina_facturacion.cliente_combo.configure(values=filtrar_nombres_clientes("", limite=200))
        self.app.pagina_facturacion.cliente_combo.set(nombre)
        self.app.pagina_facturacion.telefono.delete(0, "end")
        self.app.pagina_facturacion.telefono.insert(0, self.telefono_entry.get().strip())
        self.app.pagina_facturacion.rnc.delete(0, "end")
        self.app.pagina_facturacion.rnc.insert(0, self.rnc_entry.get().strip())
        fecha = self.fecha_entry.get().strip()
        if fecha:
            self.app.pagina_facturacion.fecha_entry.delete(0, "end")
            self.app.pagina_facturacion.fecha_entry.insert(0, fecha)

    def set_readonly_value(self, entry, value):
        entry.configure(state="normal")
        entry.delete(0, "end")
        entry.insert(0, value)
        entry.configure(state="readonly")

    def enfocar_celda_orden(self, fila_idx, col_idx):
        if not (0 <= fila_idx < len(self.filas)):
            return "break"
        if col_idx not in self.columnas_editables_orden:
            return "break"

        entry = self.filas[fila_idx][col_idx]
        entry.focus_set()
        entry.icursor("end")
        return "break"

    def mover_foco_orden(self, fila_idx, col_idx, direccion):
        columnas = self.columnas_editables_orden
        pos = columnas.index(col_idx)

        if direccion in ("next", "right"):
            if pos < len(columnas) - 1:
                return self.enfocar_celda_orden(fila_idx, columnas[pos + 1])
            if fila_idx < len(self.filas) - 1:
                return self.enfocar_celda_orden(fila_idx + 1, columnas[0])
            return "break"

        if direccion in ("prev", "left"):
            if pos > 0:
                return self.enfocar_celda_orden(fila_idx, columnas[pos - 1])
            if fila_idx > 0:
                return self.enfocar_celda_orden(fila_idx - 1, columnas[-1])
            return "break"

        if direccion == "down":
            if fila_idx < len(self.filas) - 1:
                return self.enfocar_celda_orden(fila_idx + 1, col_idx)
            return "break"

        if direccion == "up":
            if fila_idx > 0:
                return self.enfocar_celda_orden(fila_idx - 1, col_idx)
            return "break"

        return "break"

    def limpiar_resultados_fila(self, fila):
        for i in [2, 5, 6, 7, 8]:
            self.set_readonly_value(fila[i], "")

    # ================= CÁLCULO =================
    def calcular_fila(self, index):
        fila = self.filas[index]

        try:
            codigo_txt = fila[1].get().strip()

            if not codigo_txt:
                self.limpiar_resultados_fila(fila)
                self.refrescar_total_orden()
                return

            codigo = int(codigo_txt)

            tarifa = obtener_tarifa_por_codigo(codigo)
            if not tarifa:
                self.limpiar_resultados_fila(fila)
                self.refrescar_total_orden()
                return

            _, nombre, precio, extra = tarifa

            # SIEMPRE mostrar descripción al poner el código
            self.set_readonly_value(fila[2], nombre)

            cantidad_txt = fila[0].get().strip()
            ancho_txt = fila[3].get().strip()
            largo_txt = fila[4].get().strip()

            # Si todavía no han puesto medidas o cantidad, solo deja la descripción
            if not cantidad_txt or not ancho_txt or not largo_txt:
                self.set_readonly_value(fila[5], "")
                self.set_readonly_value(fila[6], f"{precio:.2f}")
                self.set_readonly_value(fila[7], "")
                self.set_readonly_value(fila[8], "")
                self.refrescar_total_orden()
                return

            cantidad = float(cantidad_txt)
            ancho = float(ancho_txt)
            largo = float(largo_txt)

            pies = calcular_pies(codigo, ancho, largo, extra)
            subtotal = pies * precio
            total = subtotal * cantidad

            self.set_readonly_value(fila[5], f"{pies:.2f}")
            self.set_readonly_value(fila[6], f"{precio:.2f}")
            self.set_readonly_value(fila[7], f"{subtotal:.2f}")
            self.set_readonly_value(fila[8], f"{total:.2f}")

        except ValueError:
            self.limpiar_resultados_fila(fila)

        self.refrescar_total_orden()

    def recalcular_todo(self):
        for i in range(len(self.filas)):
            self.calcular_fila(i)
        self.refrescar_total_orden()

    def refrescar_total_orden(self):
        total = 0.0
        for fila in self.filas:
            try:
                valor = fila[8].get().strip()
                if valor:
                    total += float(valor)
            except ValueError:
                pass

        self.total_orden_entry.delete(0, "end")
        self.total_orden_entry.insert(0, f"{total:.2f}")

    # ================= LIMPIEZA =================
    def limpiar_orden(self):
        self.cliente_entry.set("")
        self.telefono_entry.delete(0, "end")
        self.rnc_entry.delete(0, "end")
        self.fecha_entry.delete(0, "end")
        self.fecha_entry.insert(0, fecha_hoy_str())
        self.a_enmarcar_entry.delete(0, "end")
        self.notas_entry.delete(0, "end")
        self.total_orden_entry.delete(0, "end")
        self.notas_texto.delete("1.0", "end")

        for fila in self.filas:
            for entry in fila:
                if entry.cget("state") == "readonly":
                    entry.configure(state="normal")
                    entry.delete(0, "end")
                    entry.configure(state="readonly")
                else:
                    entry.delete(0, "end")

    def recolectar_datos_orden(self):
        self.recalcular_todo()

        detalles = []
        a_enmarcar = ""
        ancho_general = None
        largo_general = None

        for fila in self.filas:
            try:
                cantidad_txt = fila[0].get().strip()
                codigo_txt = fila[1].get().strip()
                descripcion_material = fila[2].get().strip()
                ancho_txt = fila[3].get().strip()
                largo_txt = fila[4].get().strip()
                pies_txt = fila[5].get().strip()
                precio_txt = fila[6].get().strip()
                subtotal_txt = fila[7].get().strip()
                total_txt = fila[8].get().strip()

                campos_editables = [cantidad_txt, codigo_txt, ancho_txt, largo_txt]
                if not any(campos_editables):
                    continue

                # Solo se exportan las filas que realmente quedaron completas para trabajar.
                # Si el usuario dejó una fila a medio llenar, simplemente se ignora.
                if not all([cantidad_txt, codigo_txt, ancho_txt, largo_txt, total_txt]):
                    continue

                if not all([
                    descripcion_material, pies_txt, precio_txt, subtotal_txt
                ]):
                    continue

                detalle = {
                    "cantidad": float(cantidad_txt),
                    "codigo_material": int(codigo_txt),
                    "descripcion_material": descripcion_material,
                    "ancho": float(ancho_txt),
                    "largo": float(largo_txt),
                    "pies": float(pies_txt),
                    "precio": float(precio_txt),
                    "subtotal": float(subtotal_txt),
                    "total": float(total_txt),
                }
                detalles.append(detalle)

                if not a_enmarcar:
                    a_enmarcar = self.a_enmarcar_entry.get().strip()
                    ancho_general = detalle["ancho"]
                    largo_general = detalle["largo"]

            except ValueError as exc:
                raise ValueError(str(exc) or "Hay valores inválidos en la orden.")

        if not detalles:
            raise ValueError("No hay filas válidas en la orden.")

        total_orden = sum(detalle["total"] for detalle in detalles)

        a_enmarcar = self.a_enmarcar_entry.get().strip()
        if not a_enmarcar:
            raise ValueError("Escribe arriba qué se va a enmarcar.")

        notas = self.notas_texto.get("1.0", "end").strip()
        if not notas:
            notas = self.notas_entry.get().strip()

        return {
            "cliente": self.cliente_entry.get().strip(),
            "telefono": self.telefono_entry.get().strip(),
            "rnc": self.rnc_entry.get().strip(),
            "fecha": self.fecha_entry.get().strip() or fecha_hoy_str(),
            "a_enmarcar": a_enmarcar,
            "notas": notas,
            "ancho": ancho_general if ancho_general is not None else 0.0,
            "largo": largo_general if largo_general is not None else 0.0,
            "total": total_orden,
            "detalles": detalles,
        }

    def asegurar_documento_desde_orden(self):
        if self.app.documento_actual_id:
            return self.app.documento_actual_id

        nombre_cliente = self.cliente_entry.get().strip()
        if not nombre_cliente:
            raise ValueError("Selecciona un cliente antes de guardar o pasar la orden.")

        cliente = obtener_cliente_por_nombre(nombre_cliente)
        if not cliente:
            raise ValueError("El cliente de la orden no existe en la base de datos.")

        tipo = self.app.pagina_facturacion.get_tipo_actual()
        doc_id = crear_documento_db(
            tipo,
            cliente[0],
            self.fecha_entry.get().strip() or fecha_hoy_str(),
        )
        self.app.documento_actual_id = doc_id

        self.app.pagina_facturacion.cargar_documento_existente(doc_id)
        return doc_id

    # ================= PASAR A FACTURACIÓN =================
    def pasar_orden_a_facturacion(self):
        for fila in self.filas:
            if fila[3].get().strip():
                break

        try:
            datos_orden = self.recolectar_datos_orden()
            doc_id = self.asegurar_documento_desde_orden()
            self.app.pagina_facturacion.registrar_orden_desde_datos(datos_orden, actualizar_vista=True)
            ruta_factura, ruta_orden = exportar_documentos(doc_id)
            self.app.pagina_facturacion.cargar_documento_existente(doc_id)
            self.app.ir_facturacion()
            self.limpiar_orden()
            messagebox.showinfo(
                "Orden guardada",
                "La orden se guardó y pasó a facturación.\n\n"
                f"Factura/Cotización: {ruta_factura or 'No se pudo generar'}\n"
                f"Orden de taller: {ruta_orden or 'No se pudo generar'}"
            )
        except Exception as exc:
            messagebox.showerror("Error", f"No se pudo pasar la orden:\n{exc}")

class PaginaFacturacion(ctk.CTkFrame):
    def __init__(self, app):
        super().__init__(app.content)
        self.app = app
        self.lineas = []
        self._recalculo_job = None
        self._cliente_values = []

        self.configure(fg_color="#F3F7F7")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # ================= ENCABEZADO =================
        header = ctk.CTkFrame(self, fg_color=COLOR_PANEL, border_width=1, border_color=COLOR_BORDE, corner_radius=18)
        header.grid(row=0, column=0, sticky="ew", padx=20, pady=(20, 10))
        header.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            header,
            text="FACTURACIÓN",
            font=ctk.CTkFont(size=30, weight="bold"),
            text_color="black"
        ).grid(row=0, column=0, padx=20, pady=(18, 6), sticky="w")

        ctk.CTkLabel(
            header,
            text="Documentos, cobros y órdenes en una sola vista",
            font=ctk.CTkFont(size=13),
            text_color=COLOR_TEXTO_SUAVE
        ).grid(row=1, column=0, padx=20, pady=(0, 18), sticky="w")

        ctk.CTkLabel(
            header,
            text="Enmarcados PF",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=COLOR_PRINCIPAL
        ).grid(row=0, column=1, rowspan=2, padx=20, pady=18, sticky="e")

        # ================= DATOS DEL CLIENTE =================
        top = ctk.CTkFrame(self, fg_color="white", border_width=2, border_color=COLOR_PRINCIPAL, corner_radius=18)
        top.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 12))

        for i in range(10):
            top.grid_columnconfigure(i, weight=1)

        ctk.CTkLabel(top, text="Tipo", text_color="black").grid(row=0, column=0, padx=8, pady=(12, 4), sticky="w")
        self.tipo_combo = ctk.CTkComboBox(
            top,
            values=["factura", "cotizacion"],
            command=self.cambiar_tipo_documento,
            fg_color="white",
            button_color=COLOR_PRINCIPAL,
            button_hover_color=COLOR_PRINCIPAL_HOVER,
            text_color="black",
            border_color=COLOR_BORDE,
            dropdown_fg_color="white",
            dropdown_text_color="black"
        )
        self.tipo_combo.grid(row=1, column=0, padx=8, pady=(0, 10), sticky="ew")
        self.tipo_combo.set("factura")

        ctk.CTkLabel(top, text="Cliente", text_color="black").grid(row=0, column=1, padx=8, pady=(12, 4), sticky="w")
        self.cliente_combo = ctk.CTkComboBox(
            top,
            values=[c[1] for c in obtener_todos_clientes()],
            command=self.seleccionar_cliente,
            fg_color="white",
            button_color=COLOR_PRINCIPAL,
            button_hover_color=COLOR_PRINCIPAL_HOVER,
            text_color="black",
            border_color=COLOR_BORDE,
            dropdown_fg_color="white",
            dropdown_text_color="black"
        )
        self.cliente_combo.grid(row=1, column=1, columnspan=2, padx=8, pady=(0, 10), sticky="ew")
        self.cliente_combo.bind("<KeyRelease>", self.buscar_cliente_facturacion)
        self._cliente_values = [c[1] for c in obtener_todos_clientes()]

        self.numero_label = ctk.CTkLabel(top, text="No. Factura", text_color="black")
        self.numero_label.grid(row=0, column=3, padx=8, pady=(12, 4), sticky="w")
        self.numero_factura = ctk.CTkEntry(
            top,
            fg_color="white",
            text_color="black",
            border_color=COLOR_BORDE
        )
        self.numero_factura.grid(row=1, column=3, padx=8, pady=(0, 10), sticky="ew")
        self.actualizar_texto_tipo()

        ctk.CTkLabel(top, text="Fecha", text_color="black").grid(row=0, column=4, padx=8, pady=(12, 4), sticky="w")
        self.fecha_entry = ctk.CTkEntry(
            top,
            fg_color="white",
            text_color="black",
            border_color=COLOR_BORDE
        )
        self.fecha_entry.grid(row=1, column=4, padx=8, pady=(0, 10), sticky="ew")
        self.fecha_entry.insert(0, fecha_hoy_str())
        ctk.CTkLabel(top, text="Teléfono", text_color="black").grid(row=0, column=5, padx=8, pady=(12, 4), sticky="w")
        self.telefono = ctk.CTkEntry(
            top,
            fg_color="white",
            text_color="black",
            border_color=COLOR_BORDE
        )
        self.telefono.grid(row=1, column=5, padx=8, pady=(0, 10), sticky="ew")

        ctk.CTkLabel(top, text="RNC", text_color="black").grid(row=0, column=6, padx=8, pady=(12, 4), sticky="w")
        self.rnc = ctk.CTkEntry(
            top,
            fg_color="white",
            text_color="black",
            border_color=COLOR_BORDE
        )
        self.rnc.grid(row=1, column=6, padx=8, pady=(0, 10), sticky="ew")

        ctk.CTkButton(
            top,
            text="Nuevo",
            fg_color=COLOR_PRINCIPAL,
            hover_color=COLOR_PRINCIPAL_HOVER,
            text_color="black",
            command=self.nuevo_documento
        ).grid(row=1, column=7, padx=8, pady=(0, 10), sticky="ew")

        ctk.CTkButton(
            top,
            text="Pasar a orden",
            fg_color="white",
            hover_color="#f2f2f2",
            text_color="black",
            border_width=1,
            border_color="black",
            command=self.ir_a_orden
        ).grid(row=1, column=8, padx=8, pady=(0, 10), sticky="ew")

        ctk.CTkLabel(top, text="Método de pago", text_color="black").grid(row=2, column=0, padx=8, pady=(2, 4), sticky="w")
        self.metodo_pago_combo = ctk.CTkComboBox(
            top,
            values=["Pendiente", "Efectivo", "Tarjeta", "Transferencia"],
            fg_color="white",
            button_color=COLOR_PRINCIPAL,
            button_hover_color=COLOR_PRINCIPAL_HOVER,
            text_color="black",
            border_color=COLOR_BORDE,
            dropdown_fg_color="white",
            dropdown_text_color="black"
        )
        self.metodo_pago_combo.grid(row=3, column=0, columnspan=2, padx=8, pady=(0, 10), sticky="ew")
        self.metodo_pago_combo.set("Pendiente")

        ctk.CTkLabel(top, text="Referencia cobro", text_color="black").grid(row=2, column=2, padx=8, pady=(2, 4), sticky="w")
        self.referencia_pago = ctk.CTkEntry(
            top,
            fg_color="white",
            text_color="black",
            border_color=COLOR_BORDE
        )
        self.referencia_pago.grid(row=3, column=2, columnspan=3, padx=8, pady=(0, 10), sticky="ew")

        # ================= TABLA FACTURA =================
        tabla_wrap = ctk.CTkFrame(self, fg_color="white", border_width=2, border_color=COLOR_PRINCIPAL, corner_radius=18)
        tabla_wrap.grid(row=2, column=0, sticky="nsew", padx=20, pady=10)
        tabla_wrap.grid_columnconfigure(0, weight=1)
        tabla_wrap.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(
            tabla_wrap,
            text="Detalle de la factura",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color="black"
        ).grid(row=0, column=0, padx=12, pady=(12, 8), sticky="w")

        self.tabla = ctk.CTkTextbox(
            tabla_wrap,
            height=360,
            fg_color="#FCFEFE",
            text_color="black",
            border_color=COLOR_BORDE,
            border_width=1
        )
        self.tabla.grid(row=1, column=0, padx=12, pady=(0, 12), sticky="nsew")
        self._reset_tabla()

        # ================= RESUMEN =================
        resumen = ctk.CTkFrame(self, fg_color="white", border_width=2, border_color=COLOR_PRINCIPAL, corner_radius=18)
        resumen.grid(row=3, column=0, sticky="ew", padx=20, pady=(0, 10))

        for i in range(6):
            resumen.grid_columnconfigure(i, weight=1)

        self.resumen_cards = {}

        def crear_card(titulo, columna, clave):
            card = ctk.CTkFrame(resumen, fg_color=COLOR_PANEL, border_width=1, border_color=COLOR_BORDE, corner_radius=14)
            card.grid(row=0, column=columna, padx=6, pady=(8, 6), sticky="ew")
            ctk.CTkLabel(card, text=titulo, font=ctk.CTkFont(size=11, weight="bold"), text_color=COLOR_TEXTO_SUAVE).pack(anchor="w", padx=10, pady=(8, 1))
            valor = ctk.CTkLabel(card, text="RD$ 0.00", font=ctk.CTkFont(size=15, weight="bold"), text_color="black")
            valor.pack(anchor="w", padx=10, pady=(0, 8))
            self.resumen_cards[clave] = valor

        crear_card("Subtotal", 0, "subtotal")
        crear_card("Descuento", 1, "descuento")
        crear_card("ITBIS", 2, "itbis")
        crear_card("Total final", 3, "total_final")
        crear_card("Abono", 4, "abono")
        crear_card("Restante", 5, "restante")

        controles = ctk.CTkFrame(resumen, fg_color="transparent")
        controles.grid(row=1, column=0, columnspan=6, sticky="ew", padx=10, pady=(0, 10))
        for i in range(4):
            controles.grid_columnconfigure(i, weight=1)

        ctk.CTkLabel(
            controles,
            text="Descuento (%)",
            text_color="black",
            font=ctk.CTkFont(size=13, weight="bold")
        ).grid(row=0, column=0, padx=(0, 8), pady=(4, 4), sticky="w")
        self.descuento = ctk.CTkEntry(
            controles,
            fg_color="white",
            text_color="black",
            border_color=COLOR_BORDE,
            width=140
        )
        self.descuento.grid(row=1, column=0, padx=(0, 12), pady=(0, 4), sticky="ew")

        ctk.CTkLabel(
            controles,
            text="Abono",
            text_color="black",
            font=ctk.CTkFont(size=13, weight="bold")
        ).grid(row=0, column=1, padx=8, pady=(4, 4), sticky="w")
        self.abono = ctk.CTkEntry(
            controles,
            fg_color="white",
            text_color="black",
            border_color=COLOR_BORDE,
            width=160
        )
        self.abono.grid(row=1, column=1, padx=8, pady=(0, 4), sticky="ew")

        ctk.CTkLabel(
            controles,
            text="Ajusta solo estos dos campos. El resto se calcula automaticamente.",
            text_color=COLOR_TEXTO_SUAVE,
            font=ctk.CTkFont(size=12)
        ).grid(row=0, column=2, columnspan=2, rowspan=2, padx=(12, 0), pady=(0, 2), sticky="e")

        # Campos internos para conservar la logica de calculo sin recargar la interfaz.
        self.subtotal = ctk.CTkEntry(self)
        self.itbis = ctk.CTkEntry(self)
        self.total_final = ctk.CTkEntry(self)
        self.restante = ctk.CTkEntry(self)
        self.actualizar_resumen_visual(calcular_resumen_financiero(0, 0, 0))

        self.descuento.bind("<KeyRelease>", self.programar_recalculo_totales)
        self.abono.bind("<KeyRelease>", self.programar_recalculo_totales)

        # ================= BOTONES =================
        botones = ctk.CTkFrame(self, fg_color="white")
        botones.grid(row=4, column=0, pady=(0, 20))

        ctk.CTkButton(
            botones,
            text="Generar PDF factura",
            fg_color="#57C7C3",
            hover_color="#3eb5b1",
            text_color="black",
            command=self.generar_pdf_factura
        ).grid(row=0, column=0, padx=10)

        ctk.CTkButton(
            botones,
            text="Imprimir factura",
            fg_color="white",
            hover_color="#f2f2f2",
            text_color="black",
            border_width=1,
            border_color="black",
            command=self.imprimir_factura
        ).grid(row=0, column=1, padx=10)

        ctk.CTkButton(
            botones,
            text="Generar PDF orden",
            fg_color="white",
            hover_color="#f2f2f2",
            text_color="black",
            border_width=1,
            border_color="black",
            command=self.generar_pdf_ordenes
        ).grid(row=0, column=2, padx=10)

        ctk.CTkButton(
            botones,
            text="Imprimir orden",
            fg_color="white",
            hover_color="#f2f2f2",
            text_color="black",
            border_width=1,
            border_color="black",
            command=self.imprimir_ordenes
        ).grid(row=0, column=3, padx=10)

        ctk.CTkButton(
            botones,
            text="Finalizar cliente",
            fg_color="black",
            hover_color="#222222",
            text_color="white",
            command=self.finalizar
        ).grid(row=0, column=4, padx=10)

        ctk.CTkButton(
            botones,
            text="Limpiar",
            fg_color="white",
            hover_color="#f2f2f2",
            text_color="black",
            border_width=1,
            border_color="black",
            command=self.nuevo_documento
        ).grid(row=0, column=5, padx=10)

    def _reset_tabla(self):
        self.tabla.configure(state="normal")
        self.tabla.delete("1.0", "end")
        self.tabla.insert("end", "Resumen de las ordenes incluidas en esta factura\n")
        self.tabla.insert("end", "=" * 88 + "\n")
        self.tabla.insert("end", f"{'#':<5}{'Cantidad':<12}{'Que se va a enmarcar':<49}{'Total':>22}\n")
        self.tabla.insert("end", "-" * 88 + "\n")
        self.tabla.configure(state="disabled")

    def _formatear_lineas_orden(self, orden, indice):
        cantidad = obtener_cantidad_principal_orden(orden)
        cantidad_txt = str(int(cantidad) if cantidad == int(cantidad) else round(cantidad, 2))
        descripcion = str(orden["a_enmarcar"]).strip() or "Sin descripcion"
        return f"{indice:<5}{cantidad_txt:<12}{descripcion[:49]:<49}{('RD$ ' + format(float(orden['total']), ',.2f')):>22}\n"

    def _set_resumen_card(self, clave, valor, prefijo="RD$ "):
        if clave in self.resumen_cards:
            self.resumen_cards[clave].configure(text=f"{prefijo}{valor}")

    def actualizar_resumen_visual(self, resumen):
        self._set_resumen_card("subtotal", f'{resumen["subtotal"]:,.2f}')
        self._set_resumen_card("descuento", f'{resumen["descuento_monto"]:,.2f} ({resumen["descuento_pct"]:.1f}%)', prefijo="")
        self._set_resumen_card("itbis", f'{resumen["itbis"]:,.2f}')
        self._set_resumen_card("total_final", f'{resumen["total_final"]:,.2f}')
        self._set_resumen_card("abono", f'{resumen["abono"]:,.2f}')
        texto_restante = f'{max(0, resumen["restante"]):,.2f}'
        if resumen["total_final"] > 0 and resumen["restante"] <= 0:
            texto_restante += " | PAGADA"
        self._set_resumen_card("restante", texto_restante)

    def programar_recalculo_totales(self, _event=None):
        if self._recalculo_job:
            self.after_cancel(self._recalculo_job)
        self._recalculo_job = self.after(140, self.calcular_totales)

    def get_tipo_actual(self):
        tipo = self.tipo_combo.get().strip().lower()
        return tipo if tipo in ("factura", "cotizacion") else "factura"

    def cambiar_tipo_documento(self, _valor=None):
        if self.app.documento_actual_id:
            messagebox.showwarning(
                "Aviso",
                "El tipo del documento actual ya está creado. Pulsa 'Nuevo' si quieres cambiar entre factura y cotización."
            )
            documento = obtener_documento(self.app.documento_actual_id)
            if documento:
                self.tipo_combo.set(documento[1])
            return

        self.actualizar_texto_tipo()
        self.numero_factura.delete(0, "end")

    def actualizar_texto_tipo(self):
        if not hasattr(self, "numero_label"):
            return
        tipo = self.get_tipo_actual()
        if tipo == "cotizacion":
            self.numero_label.configure(text="No. Cotización")
        else:
            self.numero_label.configure(text="No. Factura")

    def sync_cliente_con_orden(self):
        nombre = self.cliente_combo.get().strip()
        if nombre:
            self.app.pagina_orden.refrescar_clientes()
            self.app.pagina_orden.cliente_entry.set(nombre)
            self.app.pagina_orden.autocompletar_cliente(nombre)
            fecha = self.fecha_entry.get().strip()
            if fecha:
                self.app.pagina_orden.fecha_entry.delete(0, "end")
                self.app.pagina_orden.fecha_entry.insert(0, fecha)

    def buscar_cliente_facturacion(self, event=None):
        texto = self.cliente_combo.get().strip()
        sugerencias = filtrar_nombres_clientes(texto)
        self.cliente_combo.configure(values=sugerencias)

    def ir_a_orden(self):
        self.sync_cliente_con_orden()
        self.app.ir_orden()

    def renderizar_lineas_en_tabla(self):
        self._reset_tabla()
        self.tabla.configure(state="normal")

        contenido = "".join(
            self._formatear_lineas_orden(orden, indice)
            for indice, orden in enumerate(self.lineas, start=1)
        )
        self.tabla.insert("end", contenido)

        self.tabla.configure(state="disabled")

    def append_orden_a_tabla(self, orden):
        self.renderizar_lineas_en_tabla()

    def seleccionar_cliente(self, nombre):
        cliente = obtener_cliente_por_nombre(nombre)
        if not cliente:
            return

        cliente_id, _, telefono, rnc, _ = cliente

        self.telefono.delete(0, "end")
        self.telefono.insert(0, telefono or "")

        self.rnc.delete(0, "end")
        self.rnc.insert(0, rnc or "")
        self.sync_cliente_con_orden()

        if not self.app.documento_actual_id:
            doc_id = crear_documento_db(
                self.get_tipo_actual(),
                cliente_id,
                self.fecha_entry.get().strip() or fecha_hoy_str(),
            )
            self.app.documento_actual_id = doc_id

            documento = obtener_documento(doc_id)
            if documento:
                self.numero_factura.delete(0, "end")
                self.numero_factura.insert(0, str(documento[2]))

    def agregar_linea_visual(self, cantidad, descripcion, ancho, largo, total):
        self.lineas.append({
            "cantidad": cantidad,
            "a_enmarcar": descripcion,
            "notas": "",
            "ancho": ancho,
            "largo": largo,
            "total": total,
            "detalles": [],
        })
        self.renderizar_lineas_en_tabla()

    def registrar_orden_desde_datos(self, datos_orden, actualizar_vista=True):
        if not self.app.documento_actual_id:
            nombre = self.cliente_combo.get().strip()
            cliente = obtener_cliente_por_nombre(nombre)
            if not cliente:
                messagebox.showerror("Error", "Selecciona un cliente primero en facturación.")
                return

            doc_id = crear_documento_db(
                self.get_tipo_actual(),
                cliente[0],
                self.fecha_entry.get().strip() or fecha_hoy_str(),
            )
            self.app.documento_actual_id = doc_id

            documento = obtener_documento(doc_id)
            if documento:
                self.numero_factura.delete(0, "end")
                self.numero_factura.insert(0, str(documento[2]))

        orden_id = crear_orden_db(
            self.app.documento_actual_id,
            datos_orden["a_enmarcar"],
            datos_orden["ancho"],
            datos_orden["largo"],
            datos_orden["total"],
            datos_orden.get("notas", "")
        )

        for detalle in datos_orden["detalles"]:
            guardar_detalle_orden_db(
                orden_id,
                detalle["cantidad"],
                detalle["codigo_material"],
                detalle["descripcion_material"],
                detalle["ancho"],
                detalle["largo"],
                detalle["pies"],
                detalle["precio"],
                detalle["subtotal"],
                detalle["total"],
            )

        if actualizar_vista:
            nueva_orden = {
                "cantidad": 1,
                "a_enmarcar": datos_orden["a_enmarcar"],
                "notas": datos_orden.get("notas", ""),
                "ancho": datos_orden["ancho"],
                "largo": datos_orden["largo"],
                "total": datos_orden["total"],
                "detalles": [
                    (
                        detalle["cantidad"],
                        detalle["codigo_material"],
                        detalle["descripcion_material"],
                        detalle["ancho"],
                        detalle["largo"],
                        detalle["pies"],
                        detalle["precio"],
                        detalle["subtotal"],
                        detalle["total"],
                    )
                    for detalle in datos_orden["detalles"]
                ],
            }
            self.lineas.append(nueva_orden)
            self.append_orden_a_tabla(nueva_orden)
            self.calcular_totales()

        return orden_id

    def calcular_totales(self):
        self._recalculo_job = None
        subtotal = sum(l["total"] for l in self.lineas)

        try:
            descuento = float(self.descuento.get().strip() or 0)
        except ValueError:
            descuento = 0.0

        try:
            nuevo_abono = float(self.abono.get().strip() or 0)
        except ValueError:
            nuevo_abono = 0.0

        if self.app.documento_actual_id:
            actualizar_metodo_pago_documento(self.app.documento_actual_id, self.metodo_pago_combo.get())
            actualizar_totales_documento(self.app.documento_actual_id, descuento)
            sincronizar_cobro_automatico_facturacion(
                self.app.documento_actual_id,
                self.fecha_entry.get().strip() or fecha_hoy_str(),
                nuevo_abono,
                self.metodo_pago_combo.get(),
            )
            abono_total = obtener_total_cobrado_documento(self.app.documento_actual_id)
        else:
            abono_total = nuevo_abono

        resumen = calcular_resumen_financiero(subtotal, descuento, abono_total)

        self.subtotal.delete(0, "end")
        self.subtotal.insert(0, f'{resumen["subtotal"]:.2f}')

        self.itbis.delete(0, "end")
        self.itbis.insert(0, f'{resumen["itbis"]:.2f}')

        self.total_final.delete(0, "end")
        self.total_final.insert(0, f'{resumen["total_final"]:.2f}')

        self.restante.delete(0, "end")
        self.restante.insert(0, f'{resumen["restante"]:.2f}')
        self.actualizar_resumen_visual(resumen)

    def refrescar_resumen_desde_db(self):
        if not self.app.documento_actual_id:
            return

        documento = obtener_documento(self.app.documento_actual_id)
        if not documento:
            return

        _, _, _, _, _, _, _, _, _, subtotal, descuento, itbis, total_final, _, metodo_pago, _retirado, _fecha_entrega = documento

        auto_cobro = obtener_cobro_automatico_facturacion(self.app.documento_actual_id)
        abono_automatico = float(auto_cobro[2]) if auto_cobro else 0.0
        abono_registrado = obtener_total_cobrado_documento(self.app.documento_actual_id)
        resumen = calcular_resumen_financiero(subtotal, descuento, abono_registrado)

        self.subtotal.delete(0, "end")
        self.subtotal.insert(0, f"{subtotal:.2f}")

        self.descuento.delete(0, "end")
        self.descuento.insert(0, f"{descuento:.2f}")
        self.metodo_pago_combo.set(normalizar_metodo_pago(metodo_pago))
        self.abono.delete(0, "end")
        if abono_automatico > 0:
            self.abono.insert(0, f"{abono_automatico:.2f}")

        self.itbis.delete(0, "end")
        self.itbis.insert(0, f'{resumen["itbis"]:.2f}')

        self.total_final.delete(0, "end")
        self.total_final.insert(0, f'{resumen["total_final"]:.2f}')

        self.restante.delete(0, "end")
        self.restante.insert(0, f'{resumen["restante"]:.2f}')
        self.actualizar_resumen_visual(resumen)

    def finalizar(self):
        if not self.app.documento_actual_id:
            messagebox.showwarning("Aviso", "No hay factura activa.")
            return

        descuento = self.get_descuento()

        subtotal, descuento, itbis, total_final = actualizar_totales_documento(
            self.app.documento_actual_id,
            descuento
        )
        actualizar_metodo_pago_documento(self.app.documento_actual_id, self.metodo_pago_combo.get())
        abono_total = obtener_total_cobrado_documento(self.app.documento_actual_id)
        resumen = calcular_resumen_financiero(subtotal, descuento, abono_total)

        self.subtotal.delete(0, "end")
        self.subtotal.insert(0, f"{subtotal:.2f}")

        self.itbis.delete(0, "end")
        self.itbis.insert(0, f'{resumen["itbis"]:.2f}')

        self.total_final.delete(0, "end")
        self.total_final.insert(0, f'{resumen["total_final"]:.2f}')

        self.restante.delete(0, "end")
        self.restante.insert(0, f'{resumen["restante"]:.2f}')
        self.actualizar_resumen_visual(resumen)

        cerrar_documento_db(self.app.documento_actual_id)

        messagebox.showinfo("Guardado", "Factura finalizada y guardada en historial.")
        self.nuevo_documento()

    def generar_pdf_factura(self):
        if not self.app.documento_actual_id:
            messagebox.showwarning("Aviso", "No hay factura activa.")
            return None

        descuento = self.get_descuento()
        actualizar_totales_documento(self.app.documento_actual_id, descuento)
        try:
            monto_abono = float(self.abono.get().strip() or 0)
        except ValueError:
            monto_abono = 0.0
        sincronizar_cobro_automatico_facturacion(
            self.app.documento_actual_id,
            self.fecha_entry.get().strip() or fecha_hoy_str(),
            monto_abono,
            self.metodo_pago_combo.get(),
        )

        ruta = generar_pdf_documento(self.app.documento_actual_id)
        if ruta:
            guardar_copia_backup(ruta)
            abierto = abrir_archivo_generado(ruta)
            if abierto:
                messagebox.showinfo("PDF generado", f"Documento guardado y abierto en:\n{ruta}")
            else:
                messagebox.showwarning("PDF generado", f"Documento guardado en:\n{ruta}\n\nNo se pudo abrir automaticamente.")
            return ruta
        else:
            messagebox.showerror("Error", "No se pudo generar el PDF.")
            return None

    def imprimir_factura(self):
        if not self.app.documento_actual_id:
            messagebox.showwarning("Aviso", "No hay factura activa.")
            return
        descuento = self.get_descuento()
        actualizar_totales_documento(self.app.documento_actual_id, descuento)
        try:
            monto_abono = float(self.abono.get().strip() or 0)
        except ValueError:
            monto_abono = 0.0
        sincronizar_cobro_automatico_facturacion(
            self.app.documento_actual_id,
            self.fecha_entry.get().strip() or fecha_hoy_str(),
            monto_abono,
            self.metodo_pago_combo.get(),
        )
        ruta = generar_pdf_documento(self.app.documento_actual_id)
        if ruta:
            guardar_copia_backup(ruta)
            ok, mensaje = imprimir_archivo_generado(ruta)
            if ok:
                messagebox.showinfo("Impresión", mensaje)
            else:
                messagebox.showerror("Error de impresión", f"{mensaje}\n\nPDF disponible en:\n{ruta}")
        else:
            messagebox.showerror("Error", "No se pudo generar el PDF.")

    def generar_pdf_ordenes(self):
        if not self.app.documento_actual_id:
            messagebox.showwarning("Aviso", "No hay factura activa.")
            return None

        descuento = self.get_descuento()
        actualizar_totales_documento(self.app.documento_actual_id, descuento)
        try:
            monto_abono = float(self.abono.get().strip() or 0)
        except ValueError:
            monto_abono = 0.0
        sincronizar_cobro_automatico_facturacion(
            self.app.documento_actual_id,
            self.fecha_entry.get().strip() or fecha_hoy_str(),
            monto_abono,
            self.metodo_pago_combo.get(),
        )

        ruta = generar_pdf_ordenes(self.app.documento_actual_id)
        if ruta:
            guardar_copia_backup(ruta)
            abierto = abrir_archivo_generado(ruta)
            if abierto:
                messagebox.showinfo("PDF generado", f"Orden guardada y abierta en:\n{ruta}")
            else:
                messagebox.showwarning("PDF generado", f"Orden guardada en:\n{ruta}\n\nNo se pudo abrir automaticamente.")
            return ruta
        else:
            messagebox.showerror("Error", "No se pudo generar el PDF de la orden.")
            return None

    def imprimir_ordenes(self):
        if not self.app.documento_actual_id:
            messagebox.showwarning("Aviso", "No hay factura activa.")
            return
        descuento = self.get_descuento()
        actualizar_totales_documento(self.app.documento_actual_id, descuento)
        try:
            monto_abono = float(self.abono.get().strip() or 0)
        except ValueError:
            monto_abono = 0.0
        sincronizar_cobro_automatico_facturacion(
            self.app.documento_actual_id,
            self.fecha_entry.get().strip() or fecha_hoy_str(),
            monto_abono,
            self.metodo_pago_combo.get(),
        )
        ruta = generar_pdf_ordenes(self.app.documento_actual_id)
        if ruta:
            guardar_copia_backup(ruta)
            ok, mensaje = imprimir_archivo_generado(ruta)
            if ok:
                messagebox.showinfo("Impresión", mensaje)
            else:
                messagebox.showerror("Error de impresión", f"{mensaje}\n\nPDF disponible en:\n{ruta}")
        else:
            messagebox.showerror("Error", "No se pudo generar el PDF de la orden.")

    def get_descuento(self):
        try:
            return float(self.descuento.get().strip() or 0)
        except ValueError:
            return 0.0

    def refrescar(self):
        nuevos_clientes = filtrar_nombres_clientes("", limite=200)
        if nuevos_clientes != self._cliente_values:
            self._cliente_values = nuevos_clientes
            self.cliente_combo.configure(values=nuevos_clientes)
        self.actualizar_texto_tipo()
        self.sync_cliente_con_orden()

    def nuevo_documento(self):
        self.app.documento_actual_id = None
        self.lineas.clear()

        self.tipo_combo.set("factura")
        self.actualizar_texto_tipo()
        self.cliente_combo.set("")
        self.telefono.delete(0, "end")
        self.rnc.delete(0, "end")
        self.numero_factura.delete(0, "end")

        self.fecha_entry.delete(0, "end")
        self.fecha_entry.insert(0, fecha_hoy_str())

        self.subtotal.delete(0, "end")
        self.descuento.delete(0, "end")
        self.itbis.delete(0, "end")
        self.total_final.delete(0, "end")
        self.abono.delete(0, "end")
        self.restante.delete(0, "end")
        self.referencia_pago.delete(0, "end")
        self.metodo_pago_combo.set("Pendiente")

        self._reset_tabla()
        self.actualizar_resumen_visual(calcular_resumen_financiero(0, 0, 0))

    def cargar_documento_existente(self, documento_id):
        documento = obtener_documento(documento_id)
        if not documento:
            messagebox.showerror("Error", "Documento no encontrado.")
            return

        self.app.documento_actual_id = documento_id
        self.lineas.clear()

        _, tipo, numero_doc, _, nombre, telefono, rnc, _, fecha, subtotal, descuento, itbis, total_final, _, metodo_pago, _retirado, _fecha_entrega = documento
        auto_cobro = obtener_cobro_automatico_facturacion(documento_id)
        abono_automatico = float(auto_cobro[2]) if auto_cobro else 0.0

        self.tipo_combo.set(tipo)
        self.actualizar_texto_tipo()
        self.cliente_combo.set(nombre)

        self.telefono.delete(0, "end")
        self.telefono.insert(0, telefono or "")

        self.rnc.delete(0, "end")
        self.rnc.insert(0, rnc or "")

        self.numero_factura.delete(0, "end")
        self.numero_factura.insert(0, str(numero_doc))

        self.fecha_entry.delete(0, "end")
        self.fecha_entry.insert(0, fecha)

        self.subtotal.delete(0, "end")
        self.subtotal.insert(0, f"{subtotal:.2f}")

        self.descuento.delete(0, "end")
        self.descuento.insert(0, f"{descuento:.2f}")

        self.itbis.delete(0, "end")
        self.itbis.insert(0, f"{itbis:.2f}")

        self.total_final.delete(0, "end")
        self.total_final.insert(0, f"{total_final:.2f}")

        self.abono.delete(0, "end")
        if abono_automatico > 0:
            self.abono.insert(0, f"{abono_automatico:.2f}")
        self.restante.delete(0, "end")
        self.metodo_pago_combo.set(normalizar_metodo_pago(metodo_pago))
        self.referencia_pago.delete(0, "end")

        for orden in obtener_ordenes_completas_de_documento(documento_id):
            self.lineas.append({
                "cantidad": 1,
                "a_enmarcar": orden["a_enmarcar"],
                "notas": orden.get("notas", ""),
                "ancho": orden["ancho"],
                "largo": orden["largo"],
                "total": orden["total_orden"],
                "detalles": orden["detalles"],
            })

        self.renderizar_lineas_en_tabla()
        self.sync_cliente_con_orden()
        self.actualizar_resumen_visual(calcular_resumen_financiero(subtotal, descuento, obtener_total_cobrado_documento(documento_id)))

class PaginaHistorial(ctk.CTkFrame):
    def __init__(self, app):
        super().__init__(app.content)
        self.app = app
        self.documentos_map = {}
        self.doc_detalle_actual_id = None

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(4, weight=1)

        ctk.CTkLabel(
            self,
            text="HISTORIAL DE DOCUMENTOS",
            font=ctk.CTkFont(size=26, weight="bold")
        ).grid(row=0, column=0, padx=20, pady=(20, 10), sticky="w")

        top = ctk.CTkFrame(self)
        top.grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        top.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(top, text="Buscar:").grid(row=0, column=0, padx=8, pady=8)
        self.buscar_entry = ctk.CTkEntry(top, placeholder_text="Cliente, número, tipo o fecha")
        self.buscar_entry.grid(row=0, column=1, padx=8, pady=8, sticky="ew")
        self.buscar_entry.bind("<KeyRelease>", self.filtrar)

        ctk.CTkButton(top, text="Recargar", command=self.cargar_documentos).grid(
            row=0, column=2, padx=8, pady=8
        )

        ctk.CTkButton(top, text="Ver detalle", command=self.ver_detalle).grid(
            row=0, column=3, padx=8, pady=8
        )

        self.lista = ctk.CTkTextbox(self, height=500)
        self.lista.grid(row=2, column=0, padx=20, pady=15, sticky="nsew")

        self.id_entry = ctk.CTkEntry(self, placeholder_text="Código o ID del documento")
        self.id_entry.grid(row=3, column=0, padx=20, pady=(0, 15), sticky="ew")

        detalle_wrap = ctk.CTkFrame(self, fg_color="white", border_width=1, border_color=COLOR_BORDE, corner_radius=18)
        detalle_wrap.grid(row=4, column=0, padx=20, pady=(0, 20), sticky="nsew")
        detalle_wrap.grid_columnconfigure(0, weight=1)
        detalle_wrap.grid_rowconfigure(1, weight=1)

        acciones_detalle = ctk.CTkFrame(detalle_wrap, fg_color="transparent")
        acciones_detalle.grid(row=0, column=0, padx=12, pady=(12, 8), sticky="ew")
        for i in range(6):
            acciones_detalle.grid_columnconfigure(i, weight=1)

        self.abrir_detalle_btn = ctk.CTkButton(acciones_detalle, text="Abrir documento", command=self.abrir_documento_desde_detalle)
        self.abrir_detalle_btn.grid(row=0, column=0, padx=6, pady=4, sticky="ew")

        self.pdf_factura_btn = ctk.CTkButton(acciones_detalle, text="PDF factura", command=self.generar_pdf_documento_historial)
        self.pdf_factura_btn.grid(row=0, column=1, padx=6, pady=4, sticky="ew")

        self.imprimir_factura_btn = ctk.CTkButton(acciones_detalle, text="Imprimir factura", command=self.imprimir_documento)
        self.imprimir_factura_btn.grid(row=0, column=2, padx=6, pady=4, sticky="ew")

        self.pdf_orden_btn = ctk.CTkButton(acciones_detalle, text="PDF orden", command=self.generar_pdf_ordenes_historial)
        self.pdf_orden_btn.grid(row=0, column=3, padx=6, pady=4, sticky="ew")

        self.imprimir_orden_btn = ctk.CTkButton(acciones_detalle, text="Imprimir orden", command=self.imprimir_ordenes_historial)
        self.imprimir_orden_btn.grid(row=0, column=4, padx=6, pady=4, sticky="ew")

        self.eliminar_btn = ctk.CTkButton(
            acciones_detalle,
            text="Eliminar seleccionado",
            fg_color="#D85B5B",
            hover_color="#C14949",
            command=self.eliminar_documento_historial
        )
        self.eliminar_btn.grid(row=0, column=5, padx=6, pady=4, sticky="ew")

        self.detalle = ctk.CTkTextbox(detalle_wrap, height=260)
        self.detalle.grid(row=1, column=0, padx=12, pady=(0, 12), sticky="nsew")
        self.detalle.configure(state="disabled")

        self._set_acciones_detalle_habilitadas(False)
        self.cargar_documentos()

    def cargar_documentos(self, filtro=""):
        self.lista.configure(state="normal")
        self.lista.delete("1.0", "end")
        self.documentos_map.clear()

        filas = obtener_todos_los_documentos(filtro)

        encabezado = f"{'ID':<6} {'TIPO':<12} {'NÚMERO':<10} {'CLIENTE':<28} {'FECHA':<14} {'TOTAL':>12} {'ESTADO':<10}\n"
        self.lista.insert("end", encabezado)
        self.lista.insert("end", "-" * 105 + "\n")

        for doc_id, tipo, numero_doc, cliente, fecha, total_final, cerrado in filas:
            estado = "CERRADO" if cerrado == 1 else "ABIERTO"
            self.documentos_map[str(doc_id)] = doc_id

            linea = f"{str(doc_id):<6} {str(tipo):<12} {str(numero_doc):<10} {str(cliente)[:28]:<28} {str(fecha):<14} {float(total_final):>12.2f} {estado:<10}\n"
            self.lista.insert("end", linea)

        self.lista.configure(state="disabled")

    def filtrar(self, event=None):
        filtro = self.buscar_entry.get()
        self.cargar_documentos(filtro)

    def abrir_documento(self):
        try:
            doc_id, _documento = self.obtener_doc_id_seleccionado()
        except ValueError as exc:
            messagebox.showerror("Error", str(exc))
            return
        self.app.pagina_facturacion.cargar_documento_existente(doc_id)
        self.app.ir_facturacion()

    def abrir_documento_desde_detalle(self):
        if not self.doc_detalle_actual_id:
            messagebox.showwarning("Aviso", "Primero pulsa 'Ver detalle' en el documento que quieres abrir.")
            return
        self.app.pagina_facturacion.cargar_documento_existente(self.doc_detalle_actual_id)
        self.app.ir_facturacion()

    def _set_acciones_detalle_habilitadas(self, habilitado):
        estado = "normal" if habilitado else "disabled"
        for boton in (
            self.abrir_detalle_btn,
            self.pdf_factura_btn,
            self.imprimir_factura_btn,
            self.pdf_orden_btn,
            self.imprimir_orden_btn,
            self.eliminar_btn,
        ):
            boton.configure(state=estado)

    def obtener_doc_id_seleccionado(self):
        doc_id_txt = self.id_entry.get().strip()

        if not doc_id_txt:
            raise ValueError("Escribe el código o ID del documento.")

        try:
            identificador = int(doc_id_txt)
        except ValueError:
            raise ValueError("El código debe ser numérico.")

        documento = obtener_documento(identificador)
        if documento:
            return identificador, documento

        documento = obtener_documento_por_numero(identificador)
        if not documento:
            raise ValueError("No existe ese documento.")

        return documento[0], documento

    def ver_detalle(self):
        try:
            doc_id, documento = self.obtener_doc_id_seleccionado()
        except ValueError as exc:
            messagebox.showerror("Error", str(exc))
            return
        self.doc_detalle_actual_id = doc_id
        self._set_acciones_detalle_habilitadas(True)

        _, tipo, numero_doc, _, nombre, telefono, rnc, direccion, fecha, subtotal, descuento, itbis, total_final, cerrado, metodo_pago, retirado, fecha_entrega = documento
        ordenes = obtener_ordenes_completas_de_documento(doc_id)

        lineas = [
            f"Tipo: {tipo}",
            f"Número: {numero_doc}",
            f"Cliente: {nombre}",
            f"Teléfono: {telefono or ''}",
            f"RNC: {rnc or ''}",
            f"Dirección: {direccion or ''}",
            f"Fecha: {fecha}",
            f"Estado: {'CERRADO' if cerrado else 'ABIERTO'}",
            f"Método de pago: {normalizar_metodo_pago(metodo_pago)}",
            f"Pasó a recoger: {'SI' if retirado else 'NO'}",
            "",
            "ÓRDENES",
            "-" * 70,
        ]

        for idx, orden in enumerate(ordenes, start=1):
            lineas.append(
                f"{idx}. {orden['a_enmarcar']} | {orden['ancho']} x {orden['largo']} | Total: {orden['total_orden']:.2f}"
            )
            for detalle in orden["detalles"]:
                cantidad, codigo, descripcion_material, ancho_d, largo_d, pies, precio, subtotal_d, total_d = detalle
                lineas.append(
                    f"   - Cant: {cantidad} | Cod: {codigo} | {descripcion_material} | {ancho_d}x{largo_d} | Precio: {precio:.2f} | Total: {total_d:.2f}"
                )

        lineas.extend([
            "",
            f"Subtotal: {subtotal:.2f}",
            f"Descuento: {descuento:.2f}%",
            f"ITBIS: {itbis:.2f}",
            f"Total final: {total_final:.2f}",
            f"Cobrado: {obtener_total_cobrado_documento(doc_id):.2f}",
        ])

        self.detalle.configure(state="normal")
        self.detalle.delete("1.0", "end")
        self.detalle.insert("end", "\n".join(lineas))
        self.detalle.configure(state="disabled")

    def generar_pdf_documento_historial(self):
        try:
            doc_id, _documento = self.obtener_doc_id_seleccionado()
        except ValueError as exc:
            messagebox.showerror("Error", str(exc))
            return

        ruta = generar_pdf_documento(doc_id)
        if ruta:
            guardar_copia_backup(ruta)
            abierto = abrir_archivo_generado(ruta)
            if abierto:
                messagebox.showinfo("PDF generado", f"Documento guardado y abierto en:\n{ruta}")
            else:
                messagebox.showwarning("PDF generado", f"Documento guardado en:\n{ruta}\n\nNo se pudo abrir automaticamente.")
        else:
            messagebox.showerror("Error", "No se pudo generar el PDF.")

    def imprimir_documento(self):
        try:
            doc_id, _documento = self.obtener_doc_id_seleccionado()
        except ValueError as exc:
            messagebox.showerror("Error", str(exc))
            return

        ruta = generar_pdf_documento(doc_id)
        if ruta:
            guardar_copia_backup(ruta)
            ok, mensaje = imprimir_archivo_generado(ruta)
            if ok:
                messagebox.showinfo("Impresión", mensaje)
            else:
                messagebox.showerror("Error de impresión", f"{mensaje}\n\nPDF disponible en:\n{ruta}")
        else:
            messagebox.showerror("Error", "No se pudo generar el PDF.")

    def generar_pdf_ordenes_historial(self):
        try:
            doc_id, _documento = self.obtener_doc_id_seleccionado()
        except ValueError as exc:
            messagebox.showerror("Error", str(exc))
            return

        ruta = generar_pdf_ordenes(doc_id)
        if ruta:
            guardar_copia_backup(ruta)
            abierto = abrir_archivo_generado(ruta)
            if abierto:
                messagebox.showinfo("PDF generado", f"Orden guardada y abierta en:\n{ruta}")
            else:
                messagebox.showwarning("PDF generado", f"Orden guardada en:\n{ruta}\n\nNo se pudo abrir automaticamente.")
        else:
            messagebox.showerror("Error", "No se pudo generar el PDF de la orden.")

    def imprimir_ordenes_historial(self):
        try:
            doc_id, _documento = self.obtener_doc_id_seleccionado()
        except ValueError as exc:
            messagebox.showerror("Error", str(exc))
            return

        ruta = generar_pdf_ordenes(doc_id)
        if ruta:
            guardar_copia_backup(ruta)
            ok, mensaje = imprimir_archivo_generado(ruta)
            if ok:
                messagebox.showinfo("Impresión", mensaje)
            else:
                messagebox.showerror("Error de impresión", f"{mensaje}\n\nPDF disponible en:\n{ruta}")
        else:
            messagebox.showerror("Error", "No se pudo generar el PDF de la orden.")

    def eliminar_documento_historial(self):
        if self.doc_detalle_actual_id:
            doc_id = self.doc_detalle_actual_id
            documento = obtener_documento(doc_id)
            if not documento:
                messagebox.showerror("Error", "El documento seleccionado ya no existe.")
                return
        else:
            try:
                doc_id, documento = self.obtener_doc_id_seleccionado()
            except ValueError as exc:
                messagebox.showerror("Error", str(exc))
                return

        tipo = documento[1]
        numero = documento[2]
        cliente = documento[4]

        confirmar = messagebox.askyesno(
            "Confirmar eliminación",
            f"¿Seguro que quieres eliminar {tipo} #{numero} de {cliente}?\n\nTambién se borrarán sus órdenes y detalles."
        )
        if not confirmar:
            return

        cambios = eliminar_documento_db(doc_id)
        if cambios == 0:
            messagebox.showwarning("Aviso", "No se pudo eliminar el documento.")
            return

        if self.app.documento_actual_id == doc_id:
            self.app.pagina_facturacion.nuevo_documento()

        self.id_entry.delete(0, "end")
        self.doc_detalle_actual_id = None
        self._set_acciones_detalle_habilitadas(False)
        self.detalle.configure(state="normal")
        self.detalle.delete("1.0", "end")
        self.detalle.configure(state="disabled")
        self.cargar_documentos(self.buscar_entry.get().strip())
        messagebox.showinfo("Eliminado", "Documento eliminado correctamente.")


class PaginaCobros(ctk.CTkFrame):
    def __init__(self, app):
        super().__init__(app.content)
        self.app = app
        self.documento_actual_id = None
        self.cobro_actual_id = None

        self.configure(fg_color="#F3F7F7")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(3, weight=1)

        ctk.CTkLabel(
            self,
            text="GESTIÓN DE COBRO",
            font=ctk.CTkFont(size=26, weight="bold")
        ).grid(row=0, column=0, padx=20, pady=(20, 10), sticky="w")

        top = ctk.CTkFrame(self, fg_color="white", border_width=1, border_color=COLOR_BORDE, corner_radius=18)
        top.grid(row=1, column=0, padx=20, pady=10, sticky="ew")
        top.grid_columnconfigure(1, weight=1)
        top.grid_columnconfigure(3, weight=1)

        ctk.CTkLabel(top, text="Buscar:").grid(row=0, column=0, padx=8, pady=8)
        self.buscar_entry = ctk.CTkEntry(top, placeholder_text="Número, cliente, fecha o método de pago")
        self.buscar_entry.grid(row=0, column=1, padx=8, pady=8, sticky="ew")
        self.buscar_entry.bind("<KeyRelease>", self.filtrar)

        ctk.CTkButton(top, text="Recargar", command=self.cargar_cobros).grid(row=0, column=2, padx=8, pady=8)
        ctk.CTkLabel(top, text="ID o número:").grid(row=0, column=3, padx=(20, 8), pady=8, sticky="e")
        self.doc_entry = ctk.CTkEntry(top, placeholder_text="ID o No. factura")
        self.doc_entry.grid(row=0, column=4, padx=8, pady=8, sticky="ew")
        ctk.CTkButton(top, text="Cargar factura", command=self.cargar_factura).grid(row=0, column=5, padx=8, pady=8)

        resumen = ctk.CTkFrame(self, fg_color="white", border_width=1, border_color=COLOR_BORDE, corner_radius=18)
        resumen.grid(row=2, column=0, padx=20, pady=(0, 10), sticky="ew")
        for i in range(6):
            resumen.grid_columnconfigure(i, weight=1)

        self.info_doc = ctk.CTkLabel(resumen, text="Sin factura cargada", text_color="black", anchor="w", justify="left")
        self.info_doc.grid(row=0, column=0, columnspan=4, padx=12, pady=12, sticky="w")

        self.pagado_total_var = ctk.BooleanVar(value=False)
        self.pagado_total_check = ctk.CTkCheckBox(
            resumen,
            text="Pago completo",
            variable=self.pagado_total_var,
            command=self.aplicar_pago_total
        )
        self.pagado_total_check.grid(row=0, column=4, padx=8, pady=(0, 10), sticky="e")

        ctk.CTkLabel(resumen, text="Cobro", text_color="black").grid(row=1, column=0, padx=8, pady=(0, 4), sticky="w")
        self.monto_entry = ctk.CTkEntry(resumen, placeholder_text="Monto")
        self.monto_entry.grid(row=2, column=0, padx=8, pady=(0, 10), sticky="ew")

        ctk.CTkLabel(resumen, text="Método", text_color="black").grid(row=1, column=1, padx=8, pady=(0, 4), sticky="w")
        self.metodo_combo = ctk.CTkComboBox(resumen, values=["Pendiente", "Efectivo", "Tarjeta", "Transferencia"])
        self.metodo_combo.grid(row=2, column=1, padx=8, pady=(0, 10), sticky="ew")
        self.metodo_combo.set("Pendiente")

        ctk.CTkLabel(resumen, text="Referencia", text_color="black").grid(row=1, column=2, padx=8, pady=(0, 4), sticky="w")
        self.referencia_entry = ctk.CTkEntry(resumen, placeholder_text="Referencia")
        self.referencia_entry.grid(row=2, column=2, padx=8, pady=(0, 10), sticky="ew")

        ctk.CTkLabel(resumen, text="Fecha", text_color="black").grid(row=1, column=3, padx=8, pady=(0, 4), sticky="w")
        self.fecha_entry = ctk.CTkEntry(resumen, placeholder_text="Fecha")
        self.fecha_entry.grid(row=2, column=3, padx=8, pady=(0, 10), sticky="ew")
        self.fecha_entry.insert(0, fecha_hoy_str())

        ctk.CTkLabel(resumen, text="ID cobro", text_color="black").grid(row=1, column=4, padx=8, pady=(0, 4), sticky="w")
        self.cobro_id_entry = ctk.CTkEntry(resumen, placeholder_text="Para editar")
        self.cobro_id_entry.grid(row=2, column=4, padx=8, pady=(0, 10), sticky="ew")

        ctk.CTkButton(resumen, text="Registrar cobro", command=self.registrar_cobro).grid(row=2, column=5, padx=8, pady=(0, 10), sticky="ew")
        ctk.CTkButton(resumen, text="Editar cobro", command=self.editar_cobro).grid(row=3, column=5, padx=8, pady=(0, 10), sticky="ew")
        self.retirado_var = ctk.BooleanVar(value=False)
        self.retirado_check = ctk.CTkCheckBox(
            resumen,
            text="Pasó a recoger",
            variable=self.retirado_var,
            command=self.cambiar_estado_retirado
        )
        self.retirado_check.grid(row=3, column=4, padx=8, pady=(0, 10), sticky="w")

        cuerpo = ctk.CTkFrame(self, fg_color="transparent")
        cuerpo.grid(row=3, column=0, padx=20, pady=(0, 20), sticky="nsew")
        cuerpo.grid_columnconfigure(0, weight=1)
        cuerpo.grid_columnconfigure(1, weight=1)
        cuerpo.grid_rowconfigure(0, weight=1)

        self.lista_facturas = ctk.CTkTextbox(cuerpo, height=560, fg_color="white", text_color="black")
        self.lista_facturas.grid(row=0, column=0, padx=(0, 10), pady=0, sticky="nsew")

        self.lista_cobros = ctk.CTkTextbox(cuerpo, height=560, fg_color="white", text_color="black")
        self.lista_cobros.grid(row=0, column=1, padx=(10, 0), pady=0, sticky="nsew")

        self.cargar_cobros()

    def cargar_cobros(self, filtro=""):
        self.lista_facturas.configure(state="normal")
        self.lista_facturas.delete("1.0", "end")
        self.lista_cobros.configure(state="normal")
        self.lista_cobros.delete("1.0", "end")

        docs = obtener_resumen_documentos_cobro(filtro)
        self.lista_facturas.insert("end", "Facturas encontradas\n")
        self.lista_facturas.insert("end", "=" * 92 + "\n")
        for doc_id, numero_doc, cliente, fecha, total_final, metodo_pago, retirado, cobrado in docs:
            restante = max(0.0, float(total_final) - float(cobrado))
            retiro_txt = "SI" if retirado else "NO"
            self.lista_facturas.insert(
                "end",
                f"ID {doc_id} | Doc {numero_doc} | {str(cliente)[:28]}\n"
                f"Fecha: {fecha} | Total: {float(total_final):,.2f} | Abonos: {float(cobrado):,.2f} | Resta: {restante:,.2f}\n"
                f"Pago: {str(metodo_pago)} | Recogido: {retiro_txt}\n"
                + "-" * 92 + "\n"
            )

        self.lista_cobros.insert("end", "Carga una factura para ver su detalle completo y sus cobros.\n")

        self.lista_facturas.configure(state="disabled")
        self.lista_cobros.configure(state="disabled")

    def filtrar(self, event=None):
        self.cargar_cobros(self.buscar_entry.get().strip())

    def obtener_documento_desde_campo(self):
        codigo = self.doc_entry.get().strip()
        if not codigo:
            raise ValueError("Escribe el ID o el número de factura.")
        try:
            identificador = int(codigo)
        except ValueError:
            raise ValueError("El ID o número debe ser numérico.")

        documento = obtener_documento(identificador)
        if documento:
            return documento

        documento = obtener_documento_por_numero(identificador)
        if documento:
            return documento

        raise ValueError("No existe una factura con ese ID o número.")

    def actualizar_detalle_documento(self, documento):
        doc_id = documento[0]
        _, tipo, numero_doc, _, nombre, telefono, rnc, direccion, fecha, subtotal, descuento, _itbis, total_final, cerrado, metodo_pago, retirado, fecha_entrega = documento
        cobrado = obtener_total_cobrado_documento(doc_id)
        restante = max(0.0, float(total_final) - cobrado)
        estado = "PAGADA" if restante <= 0 and total_final > 0 else "PENDIENTE"
        ordenes = obtener_ordenes_completas_de_documento(doc_id)
        cobros = obtener_cobros_documento(doc_id)

        resumen = [
            "DETALLE DE COBRO",
            "=" * 92,
            "",
            f"Documento: {tipo.upper()} #{numero_doc}",
            f"ID interno: {doc_id}",
            "",
            "Cliente",
            f"  Nombre: {nombre}",
            f"  Telefono: {telefono or '-'}",
            f"  RNC: {rnc or '-'}",
            f"  Direccion: {direccion or '-'}",
            "",
            "Estado del documento",
            f"  Fecha: {fecha}",
            f"  Estado general: {'CERRADO' if cerrado else 'ABIERTO'}",
            f"  Recogido: {'SI' if retirado else 'NO'}",
            "",
            "Resumen de pago",
            f"  Subtotal: RD$ {subtotal:,.2f}",
            f"  Descuento: {descuento:.2f}%",
            f"  Total factura: RD$ {total_final:,.2f}",
            f"  Abonos registrados: RD$ {cobrado:,.2f}",
            f"  Restante: RD$ {restante:,.2f}",
            f"  Metodo actual: {normalizar_metodo_pago(metodo_pago)}",
            f"  Estado de cobro: {estado}",
            "",
            "Ordenes incluidas",
            "-" * 92,
        ]
        for idx, orden in enumerate(ordenes, start=1):
            resumen.append(
                f"{idx}. Cantidad: {obtener_cantidad_principal_orden(orden):g}"
            )
            resumen.append(f"   Trabajo: {orden['a_enmarcar']}")
            resumen.append(f"   Total: RD$ {float(orden['total_orden']):,.2f}")
            resumen.append("")

        resumen.extend(["Cobros registrados", "-" * 92])
        if cobros:
            for cobro_id, fecha_cobro, monto_cobro, metodo_cobro, referencia, pagado_total in cobros:
                estado_cobro = "PAGADA" if pagado_total else "PARCIAL"
                resumen.append(f"ID cobro: {cobro_id}")
                resumen.append(f"  Fecha: {fecha_cobro}")
                resumen.append(f"  Monto: RD$ {float(monto_cobro):,.2f}")
                resumen.append(f"  Metodo: {metodo_cobro}")
                resumen.append(f"  Estado: {estado_cobro}")
                resumen.append(f"  Referencia: {referencia or '-'}")
                resumen.append("")
        else:
            resumen.append("Sin cobros registrados.")

        self.lista_cobros.configure(state="normal")
        self.lista_cobros.delete("1.0", "end")
        self.lista_cobros.insert("end", "\n".join(resumen))
        self.lista_cobros.configure(state="disabled")

    def cargar_factura(self):
        try:
            documento = self.obtener_documento_desde_campo()
        except ValueError as exc:
            messagebox.showerror("Error", str(exc))
            return

        self.documento_actual_id = documento[0]
        _, tipo, numero_doc, _, nombre, _telefono, _rnc, _direccion, fecha, subtotal, descuento, _itbis, total_final, _cerrado, metodo_pago, retirado, fecha_entrega = documento
        cobrado = obtener_total_cobrado_documento(self.documento_actual_id)
        restante = max(0.0, float(total_final) - cobrado)
        estado = "PAGADA" if restante <= 0 and total_final > 0 else "PENDIENTE"
        self.info_doc.configure(
            text=(
                f"Documento: {tipo} #{numero_doc}\n"
                f"Cliente: {nombre}\n"
                f"Subtotal: {subtotal:.2f} | Descuento: {descuento:.2f}% | Total: {total_final:.2f}\n"
                f"Abonos: {cobrado:.2f} | Restante: {restante:.2f} | Estado: {estado}"
            )
        )
        self.metodo_combo.set(normalizar_metodo_pago(metodo_pago))
        self.retirado_var.set(bool(retirado))
        self.pagado_total_var.set(restante <= 0 and total_final > 0)
        self.fecha_entry.delete(0, "end")
        self.fecha_entry.insert(0, fecha_hoy_str())
        self.monto_entry.delete(0, "end")
        self.referencia_entry.delete(0, "end")
        self.cobro_actual_id = None
        self.cobro_id_entry.delete(0, "end")
        self.actualizar_detalle_documento(documento)

    def aplicar_pago_total(self):
        if not self.documento_actual_id or not self.pagado_total_var.get():
            return
        documento = obtener_documento(self.documento_actual_id)
        if not documento:
            return
        total_final = float(documento[12] or 0)
        cobrado = obtener_total_cobrado_documento(self.documento_actual_id)
        restante = max(0.0, total_final - cobrado)
        self.monto_entry.delete(0, "end")
        self.monto_entry.insert(0, f"{restante:.2f}")

    def cambiar_estado_retirado(self):
        if not self.documento_actual_id:
            return
        actualizar_estado_retiro_documento(self.documento_actual_id, self.retirado_var.get())
        documento = obtener_documento(self.documento_actual_id)
        if documento:
            self.actualizar_detalle_documento(documento)
            _, tipo, numero_doc, _, nombre, _telefono, _rnc, _direccion, _fecha, subtotal, descuento, _itbis, total_final, _cerrado, _metodo_pago, retirado, _fecha_entrega = documento
            cobrado = obtener_total_cobrado_documento(self.documento_actual_id)
            restante = max(0.0, float(total_final) - cobrado)
            estado = "PAGADA" if restante <= 0 and total_final > 0 else "PENDIENTE"
            self.info_doc.configure(
                text=(
                    f"Documento: {tipo} #{numero_doc}\n"
                    f"Cliente: {nombre}\n"
                    f"Subtotal: {subtotal:.2f} | Descuento: {descuento:.2f}% | Total: {total_final:.2f}\n"
                    f"Abonos: {cobrado:.2f} | Restante: {restante:.2f} | Estado: {estado} | Recogido: {'SI' if retirado else 'NO'}"
                )
            )
        self.cargar_cobros(self.buscar_entry.get().strip())

    def registrar_cobro(self):
        if not self.documento_actual_id:
            messagebox.showwarning("Aviso", "Primero carga una factura.")
            return
        try:
            monto = float(self.monto_entry.get().strip() or 0)
        except ValueError:
            messagebox.showerror("Error", "El monto debe ser numérico.")
            return
        if monto <= 0:
            messagebox.showwarning("Aviso", "El cobro debe ser mayor que cero.")
            return

        registrar_cobro_db(
            self.documento_actual_id,
            self.fecha_entry.get().strip() or fecha_hoy_str(),
            monto,
            self.metodo_combo.get(),
            self.referencia_entry.get().strip(),
        )
        actualizar_estado_retiro_documento(self.documento_actual_id, self.retirado_var.get())
        self.cargar_factura()
        self.cargar_cobros(self.buscar_entry.get().strip())
        self.cobro_id_entry.delete(0, "end")
        messagebox.showinfo("Cobro", "Cobro registrado correctamente.")

    def editar_cobro(self):
        if not self.cobro_actual_id:
            cobro_id_txt = self.cobro_id_entry.get().strip()
            if cobro_id_txt.isdigit():
                self.cobro_actual_id = int(cobro_id_txt)
        if not self.cobro_actual_id:
            messagebox.showwarning("Aviso", "Escribe el ID del cobro para editarlo.")
            return

        try:
            monto = float(self.monto_entry.get().strip() or 0)
        except ValueError:
            messagebox.showerror("Error", "El monto debe ser numérico.")
            return

        cambios = actualizar_cobro_db(
            self.cobro_actual_id,
            self.fecha_entry.get().strip() or fecha_hoy_str(),
            monto,
            self.metodo_combo.get(),
            self.referencia_entry.get().strip(),
        )
        if cambios == 0:
            messagebox.showwarning("Aviso", "No se pudo editar el cobro.")
            return

        cobro = obtener_cobro_por_id(self.cobro_actual_id)
        if cobro:
            actualizar_estado_retiro_documento(cobro[1], self.retirado_var.get())
            self.documento_actual_id = cobro[1]
        self.cargar_factura()
        self.cargar_cobros(self.buscar_entry.get().strip())
        self.cobro_id_entry.delete(0, "end")
        messagebox.showinfo("Cobro", "Cobro actualizado correctamente.")


class PaginaBackup(ctk.CTkFrame):
    def __init__(self, app):
        super().__init__(app.content)
        self.app = app

        self.configure(fg_color="#F3F7F7")
        self.grid_columnconfigure(0, weight=1)

        header = ctk.CTkFrame(self, fg_color=COLOR_PANEL, border_width=1, border_color=COLOR_BORDE, corner_radius=18)
        header.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="ew")
        header.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            header,
            text="BACKUP Y LIMPIEZA",
            font=ctk.CTkFont(size=26, weight="bold")
        ).grid(row=0, column=0, padx=20, pady=(16, 4), sticky="w")

        ctk.CTkLabel(
            header,
            text="Guarda los PDFs y la base actual en una carpeta con fecha, limpia facturas viejas y reinicia la numeracion.",
            font=ctk.CTkFont(size=13),
            text_color=COLOR_TEXTO_SUAVE
        ).grid(row=1, column=0, padx=20, pady=(0, 16), sticky="w")

        cuerpo = ctk.CTkFrame(self, fg_color="white", border_width=2, border_color=COLOR_PRINCIPAL, corner_radius=18)
        cuerpo.grid(row=1, column=0, padx=20, pady=(0, 20), sticky="ew")
        cuerpo.grid_columnconfigure(0, weight=1)

        texto = (
            "Este proceso hace lo siguiente:\n\n"
            "1. Crea una carpeta de respaldo con la fecha y la hora.\n"
            "2. Guarda una copia de la base de datos actual.\n"
            "3. Mueve los PDFs de 'documentos_generados' y 'facturaspf realizadas'.\n"
            "4. Borra documentos, ordenes, detalles y cobros.\n"
            "5. Reinicia los ID y la numeracion de facturas.\n\n"
            "Si estas en modo nube, tambien guarda un respaldo JSON de Railway y limpia los documentos en Railway.\n\n"
            "Clientes y tarifas se conservan."
        )
        ctk.CTkLabel(
            cuerpo,
            text=texto,
            justify="left",
            anchor="w",
            text_color="black",
            font=ctk.CTkFont(size=15)
        ).grid(row=0, column=0, padx=20, pady=(20, 16), sticky="w")

        ctk.CTkButton(
            cuerpo,
            text="Crear backup y reiniciar facturacion",
            fg_color="#D85B5B",
            hover_color="#C14949",
            text_color="white",
            command=self.ejecutar_backup
        ).grid(row=1, column=0, padx=20, pady=(0, 20), sticky="w")

    def ejecutar_backup(self):
        confirmar = messagebox.askyesno(
            "Confirmar backup",
            "Se moveran los archivos actuales al backup y se borraran documentos, ordenes y cobros del sistema.\n"
            "Si estas en modo nube, tambien se eliminaran en Railway.\n\n¿Deseas continuar?"
        )
        if not confirmar:
            return

        try:
            resultado = crear_backup_y_reiniciar_facturacion()
        except Exception as exc:
            messagebox.showerror("Error", f"No se pudo crear el backup:\n{exc}")
            return

        self.app.documento_actual_id = None
        self.app.pagina_facturacion.nuevo_documento()
        self.app.pagina_orden.limpiar_orden()
        self.app.pagina_historial.cargar_documentos()
        self.app.pagina_cobros.cargar_cobros()

        messagebox.showinfo(
            "Backup creado",
            "Se completo el respaldo correctamente.\n\n"
            f"Carpeta: {resultado['carpeta']}\n"
            f"PDFs movidos desde documentos_generados: {resultado['documentos_generados']}\n"
            f"PDFs movidos desde facturaspf realizadas: {resultado['respaldos_pdf']}\n"
            f"Documentos borrados en Railway: {resultado['borrados_nube']}"
        )

if __name__ == "__main__":
    app = App()
    app.mainloop()
    app.py
